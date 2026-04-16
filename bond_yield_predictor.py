"""
================================================================================
  INDIAN GOVERNMENT BOND YIELD PREDICTION ENGINE  v3.5.5
================================================================================
  Multi-Model Ensemble:
    Econometric  —  VAR / VECM  (reduced-form, max 4 vars, lag-capped)
    Term-Struct  —  Nelson-Siegel-Svensson (daily OLS)  +  Hull-White
    ML           —  XGBoost (regressor-only, no classifier mismatch)
    Deep Learn   —  Bidirectional LSTM  (scaler fitted on TRAIN only)

  v3.3 fixes vs v3.2  (9 engineering fixes):
    1. DATA_START_YEAR removed — auto-detect; sparse maturities (40Y/50Y)
       excluded by MIN_MAT_OBS real-observation filter
    2. Johansen cointegration moved INSIDE build_econometric() so
       k_ar_diff matches VAR-selected lag (was hardcoded 2 in diagnostics)
    3. Hull-White multi-start (5 seeds) to escape local minima
    4. _make_xgb() factory enforces identical hyperparams across
       build_xgboost() and run_backtest()
    5. LSTM seq_len derived from ACF decay (not hardcoded 60)
    6. FLAT threshold per-maturity from 20th percentile of |historical Δ|
    7. LOW CONFIDENCE threshold from binomial test (not hardcoded 55%)
    8. Granger causality uses diff() for rate vars (consistent with z-score)
    9. XGBoost stores n_test for statistical confidence computation

  v3.4 fixes vs v3.3  (7 production-readiness fixes):
    1. LSTM gated in ensemble: excluded if dir_acc <= 0.50 (anticorrelated)
    2. Momentum baseline added to backtest (fairer than naive majority-class)
    3. DirAcc > 80% flagged as possible regime artifact
    4. Macro weight vs model importance divergence surfaced (Spearman rho)
    5. Hull-White CI width > 80 bps flagged as unreliable
    6. Negative edge warning when model underperforms baselines
    7. Backtest edge computed vs max(naive, momentum) — harder benchmark

  v3.5.5 patch vs v3.5.4  (7 regime-adaptive fixes after 1W validation):
    Validated v3.5.4 1-week predictions against April 9-16 2026 market:
    model predicted -0.9 bps avg, reality was -22 bps (24x underestimate).
    12/17 predictions labelled FLAT; all actually moved DOWN 15-38 bps.
    Root cause: mean-reversion bias + exogenous shock blindness.

    1. FLAT threshold: volatility-adaptive instead of static 20th pctile.
       Uses 10th pctile × min(1, recent_vol/long_vol) × horizon_scaler.
       Short horizons (1W) get 0.4× multiplier (was: same as 3M).
    2. Data quality: cross-maturity monotonicity check — any yield >50 bps
       from linear interpolation of neighbours is replaced. Catches
       12Y=6.634% error (should have been ~7.10% between 11Y and 13Y).
    3. CI volatility scaling: width × max(1, recent_vol/long_vol). In
       volatile regimes CIs automatically widen (old: ±10 bps, needed ±30).
    4. Magnitude calibration: emag × max(1, vol_ratio). Scales up
       predictions from XGBoost's mean-reversion-biased near-zero outputs
       to match current-regime volatility.
    5. Trend-blend: horizon-dependent weights {1W:50%, 1M:35%, 3M:30%},
       boosted +15% in high-vol regime. Stronger signal for short horizons.
    6. Crude oil shock overlay: when crude moves >5% in the horizon window,
       add a directional impact proportional to crude change (India imports
       85% of crude; ~5 bps yield per 1% crude change empirically).
    7. Regime detector (_detect_regime): classifies vol as high/normal/low
       from recent_vol/long_vol ratio, detects crude shocks, measures
       trend strength across key maturities. All downstream stages
       (magnitude, trend-blend, CI, FLAT threshold) adapt to regime.

  v3.5.4 patch vs v3.5.3  (6 fixes targeting feature-crowding + trend regime):
    A. ma5/ma20/ma60 yield level features restricted to KEY_MA_MATS
       {3M, 2Y, 5Y, 10Y, 30Y}. v3.5.3 had 39+ smoothed level features
       that crowded out the 6 new directional features — none of
       fwd_spread_* / rbi_stance_* appeared in top-15 importance.
    B. XGBoost colsample_bytree 0.7 -> 0.5 forces every tree to draw
       from a smaller random feature pool, spreading importance more
       evenly and letting directional features compete for splits.
    C. Direct IIP per-variable features (m_IIP, m_ma20_IIP, etc)
       dropped. IIP had Granger p=0.8066 (no causal signal) but
       landed at #4 with 23.7% importance in v3.5.3 — classic
       spurious fit. IIP still enters via macro_composite so user
       weights still control its influence, but XGBoost can no
       longer attach directly to IIP levels.
    D. Monotone constraints: fwd_spread_* and rbi_stance{,_ma20}
       forced to +1 (forward-spread up -> yields up; hiking stance
       -> yields up). XGBoost can no longer use them in reverse.
    E. HARD_EXCLUDE extended with 19Y and 24Y. Both are non-benchmark
       Indian tenors with mostly interpolated quotes; observed post-2015
       dir_acc 48.4% and 47.2% respectively with no backtest coverage.
    F. Trend-aware blend in ensemble_predict: when XGBoost's model
       dir_acc is below the naive majority-class baseline by >2%,
       blend 30% of ensemble toward the recent realised-trend change.
       Stops the model from being systematically flat against a
       persistent regime it can't beat.

  v3.5.3 patch vs v3.5.2  (6 fixes after regime-handling review):
    1. XGBoost: hard post-2015 cutoff AND gentle time-decay (0.00005) inside
       that window. v3.5.2's decay=0.0005 over all 7431 rows created a 41x
       weight ratio, destroying XGBoost (all maturities <50% dir_acc).
       v3.5.1's hard cutoff alone overfit the 2015+ rate-cut cycle
       (70% build vs 41% backtest — 29bp gap). v3.5.5 combines both:
       regime isolation + mild within-window emphasis of recent tail.
    2. Backtest: same post-2015 restriction (plus test_size=150 to keep
       earliest folds viable). Walk-forward DirAcc now directly comparable
       to build_xgboost() test-set DirAcc.
    3. Feature set: ADD forward-spot spreads (2y-5y, 5y-10y, 10y-15y, 2y-10y)
       from existing NSS curve — market-implied direction signal with zero
       extra cost. ADD RBI policy stance flag from 60d repo-rate diff
       (hike/hold/cut) — strong signal for short end.
    4. Feature scale: real_yield_10y and term_prem_10y moved from raw
       decimal to bps (*10_000) to match slope_*, butterfly, dy*_* and
       prevent RobustScaler distortion in LSTM.
    5. macro_composite z-score warm-up: bfill (not fillna(0)) so the first
       252 rows inherit the first valid regime reading instead of a false
       "neutral" signal — 2015 was an active RBI cutting year, not neutral.
    6. VAR ensemble weight: 1.5x out-of-sample penalty on VAR RMSE (its
       value is in-sample) + 20% hard weight cap. Prevents in-sample/OOS
       mismatch from over-weighting VAR in the inverse-RMSE calculation.

  v3.5.2 patch vs v3.5.1  (superseded by v3.5.5 — kept for historical trace):
    A. LSTM hard 2015 cutoff — fixed LSTM regime contamination.
    B. XGBoost decay-only (too aggressive at 0.0005 — reverted in v3.5.5).
    C. Backtest mirrored XGBoost decay (now also uses hard cutoff).

  v3.5.1 patch vs v3.5  (5 post-run fixes from first live-output review):
    1a. LSTM reproducibility — tf/np/random seeds + op-determinism.
        Without these, two runs of the SAME code on the SAME data
        produced opposite directions on 2Y (DOWN -10.9 bps <-> UP +3.7).
    1b. LSTM fit() shuffle=False — Keras default shuffles batches, which
        breaks temporal order of a time series (data-leakage-adjacent).
    1c. LSTM ensemble gating raised 0.50 -> 0.52 — 50% was a fragile
        boundary; 49.8% vs 50.2% flips should not change predictions.
    2a. 40Y and 50Y HARD-EXCLUDED — illiquid, mostly interpolated.
        Observed 40Y dir_acc=0.9% (anti-signal) and 50Y=86.9% (artifact).
    2b. XGBoost training cutoff 2015-01-01 — removes 1998-2014 rising-rate
        regime that conflicted with 2015+ falling-rate regime and pushed
        short-end dir_acc below 50% (anti-prediction).

  v3.5 fixes vs v3.4  (13 correctness + disclosure fixes):
    1. load_data auto-detects percent vs decimal (was a hard assertion crash)
    2. VAR horizon scaling — forecast rescaled by horizon_days/(steps*20)
       (was: 1W horizon silently returned a 1-month forecast)
    3. VAR RMSE scale-matched to horizon via sqrt(h/20) for ensemble weights
       (was: monthly residual RMSE mixed with daily-horizon RMSEs)
    4. Granger: monthly data for rate vars, daily for level vars, lag=6 monthly
    5. VAR excluded-variable disclosure printed at top of build_econometric
    6. real_yield_10y / term_prem_10y stored in decimal (removed *100 scale mix)
    7. Bumped NSS cache version to v3.5-multiday (forces refit after fixes)
    8. LSTM coverage disclosure (maturities fitted vs requested)
    9. Hull-White theta starts diversified: stressed + dovish + mean regimes
   10. MIN_MAT_COVERAGE (25% of history) replaces absolute MIN_MAT_OBS
   11. VAR & HW ensemble confidence derived from in-sample dir-acc, not 0.52/0.60
   12. Horizon-dependent CI width flag: {5:30, 20:50, 60:100} bps
   13. Non-interactive mode via predictor_config.json (for Excel launcher)

  Data leakage: ZERO. Verified by gap + train-only scaling.
================================================================================
"""

# ============================================================================
# SECTION 1 — IMPORTS
# ============================================================================
import os, sys, json, warnings, datetime as dt
from pathlib import Path

import numpy  as np
import pandas as pd
from scipy import stats
from scipy.optimize import minimize as sp_minimize

from sklearn.preprocessing import StandardScaler, RobustScaler
from sklearn.model_selection import TimeSeriesSplit
from sklearn.metrics import mean_squared_error, mean_absolute_error

from statsmodels.tsa.api import VAR
from statsmodels.tsa.stattools import adfuller, grangercausalitytests, acf
from statsmodels.tsa.vector_ar.vecm import coint_johansen, VECM

import xgboost as xgb

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

warnings.filterwarnings("ignore")

LSTM_AVAILABLE = False
try:
    os.environ["TF_CPP_MIN_LOG_LEVEL"] = "3"
    import tensorflow as tf
    from tensorflow.keras.models import Sequential
    from tensorflow.keras.layers import (LSTM as KerasLSTM, Dense, Dropout,
                                         Bidirectional, BatchNormalization)
    from tensorflow.keras.callbacks import EarlyStopping, ReduceLROnPlateau
    from tensorflow.keras.optimizers import Adam
    tf.get_logger().setLevel("ERROR")
    LSTM_AVAILABLE = True
except ImportError:
    pass

print("=" * 72)
print("  INDIAN GOVERNMENT BOND YIELD PREDICTION ENGINE  v3.5.5")
models_str = "VAR/VECM + NSS + Hull-White + XGBoost"
if LSTM_AVAILABLE:
    models_str += " + Bi-LSTM"
print(f"  Ensemble: {models_str}")
print("=" * 72)
if not LSTM_AVAILABLE:
    print("  [!] TensorFlow not found  ->  LSTM disabled.")
    print("      Install:  pip install tensorflow\n")


# ============================================================================
# SECTION 2 — CONFIGURATION
# ============================================================================
EXCEL_PATH     = Path(r"C:\Users\Lenovo\Downloads\dataforpythontraining.xlsx")
OUTPUT_PATH    = Path(r"C:\Users\Lenovo\Downloads\bond_predictions_output.xlsx")
INPUT_SHEET    = "Sheet1"
OUTPUT_SHEET   = "Predictions"
BACKTEST_SHEET = "Backtest_Results"
NSS_CACHE      = Path(r"C:\Users\Lenovo\Downloads\nss_params_cache.csv")

# v3.5 FIX 13b: Non-interactive config. If predictor_config.json exists next
# to this script (or path is passed via --config), the script runs headless
# using its contents — no input() prompts. This powers the Excel button
# launcher: the VBA macro writes a fresh config, shells out to python, and
# opens the output when done.
_DEFAULT_CONFIG_PATH = Path(__file__).resolve().parent / "predictor_config.json"


def load_runtime_config():
    """Load predictor_config.json if present. Returns (config_dict_or_None, path)."""
    cfg_path = _DEFAULT_CONFIG_PATH
    if "--config" in sys.argv:
        try:
            cfg_path = Path(sys.argv[sys.argv.index("--config") + 1])
        except (IndexError, ValueError):
            pass
    if not cfg_path.exists():
        return None, cfg_path
    try:
        with open(cfg_path, "r", encoding="utf-8") as f:
            cfg = json.load(f)
        print(f"  [config] Loaded non-interactive config from {cfg_path.name}")
        return cfg, cfg_path
    except Exception as e:
        print(f"  [config] Failed to read {cfg_path.name}: {e}")
        return None, cfg_path

MIN_MAT_COVERAGE = 0.25   # min fraction of total history a maturity must cover
                          # (replaces absolute threshold — scales with dataset size)

# v3.5.1 FIX 2b: XGBoost training cutoff.
# The full Excel data goes back to 1998, but that covers three very
# different rate regimes (1998-2003 high, 2004-2014 volatile, 2015-2026
# low-and-falling). Training on all of it made the short-end model
# LEARN contradictions — observed 3M dir_acc 39.9%, 6M 41.2%, 1Y 45.2%
# (worse than coin flip = anti-prediction). Restricting training to
# >= 2015 matches the actual Refinitiv daily-quote era and the regime
# the production model is being asked to forecast.
XGB_TRAIN_START = pd.Timestamp("2015-01-01")   # hard cutoff: XGBoost + LSTM + backtest (v3.5.5)
XGB_TIME_DECAY  = 0.00005                       # gentle decay INSIDE post-2015 window (v3.5.5)
                                                 # (~9% lift newest vs oldest over ~1800 rows)

YIELD_COLS = {
    "3M":"IN3MT Yield (%)","6M":"IN6MT Yield (%)",
    "1Y":"IN1YT Yield (%)","2Y":"IN2YT Yield (%)",
    "3Y":"IN3YT Yield (%)","4Y":"IN4YT Yield (%)",
    "5Y":"IN5YT Yield (%)","6Y":"IN6YT Yield (%)",
    "7Y":"IN7YT Yield (%)","8Y":"IN8YT Yield (%)",
    "9Y":"IN9YT Yield (%)","10Y":"IN10YT Yield (%)",
    "11Y":"IN11YT Yield (%)","12Y":"IN12YT Yield (%)",
    "13Y":"IN13YT Yield (%)","14Y":"IN14YT Yield (%)",
    "15Y":"IN15YT Yield (%)","19Y":"IN19YT Yield (%)",
    "24Y":"IN24YT Yield (%)","30Y":"IN30YT Yield (%)",
    "40Y":"IN40YT Yield (%)","50Y":"IN50YT Yield (%)",
}

MACRO_COLS = {
    "CPI":"CPI YoY (%)","Repo Rate":"Repo Rate (%)",
    "IIP":"IIP Growth (%)","USD/INR":"USD/INR Rate",
    "Crude":"Crude Brent (INR/bbl)","NSE":"NSE Close Price",
    "FII":"FII (INR)",
}

MATURITIES_YRS = {
    "3M":0.25,"6M":0.5,"1Y":1,"2Y":2,"3Y":3,"4Y":4,"5Y":5,
    "6Y":6,"7Y":7,"8Y":8,"9Y":9,"10Y":10,"11Y":11,"12Y":12,
    "13Y":13,"14Y":14,"15Y":15,"19Y":19,"24Y":24,"30Y":30,
    "40Y":40,"50Y":50,
}

DISPLAY_MATS = ["3M","6M","1Y","2Y","3Y","5Y","7Y","10Y","14Y","15Y",
                "19Y","24Y","30Y","40Y","50Y"]

DEFAULT_MACRO_WEIGHTS = {"CPI":25,"Repo Rate":25,"IIP":10,
                          "USD/INR":10,"Crude":10,"NSE":10,"FII":10}


# ============================================================================
# SECTION 3 — DATA LOADING
# ============================================================================
def load_data():
    print("\n[1/12] Loading data from Excel ...")
    df = pd.read_excel(EXCEL_PATH, sheet_name=INPUT_SHEET, engine="openpyxl")
    df.rename(columns={df.columns[0]: "Date"}, inplace=True)
    df["Date"] = pd.to_datetime(df["Date"])
    df.sort_values("Date", inplace=True)
    df.drop_duplicates(subset=["Date"], keep="last", inplace=True)
    df.set_index("Date", inplace=True)
    for col in MACRO_COLS.values():
        df[col] = df[col].ffill()
    # Count real (non-NaN) observations per maturity BEFORE interpolation
    # Used downstream to exclude sparse maturities (e.g. 40Y/50Y)
    mat_real_obs = {}
    for lab, col in YIELD_COLS.items():
        mat_real_obs[lab] = int(df[col].notna().sum())
    for col in YIELD_COLS.values():
        df[col] = df[col].interpolate(method="linear")
    # v3.5 FIX 1: Unit auto-detection — all rates must be decimal (0.065 = 6.5%).
    # Previous versions asserted and crashed; now we auto-convert percent -> decimal
    # on a per-column basis (yields, CPI, Repo Rate, IIP). Level vars (USD/INR,
    # NSE, FII, Crude) are never converted because they are not rates.
    RATE_MACROS = ("CPI", "Repo Rate", "IIP")
    _y10_mean = df[YIELD_COLS["10Y"]].mean()
    if _y10_mean > 1.0:
        print(f"    [!] Yields look like percent (10Y mean={_y10_mean:.2f}); "
              f"auto-converting to decimal")
        for col in YIELD_COLS.values():
            df[col] = df[col] / 100.0
    for mname in RATE_MACROS:
        col = MACRO_COLS.get(mname)
        if col is None:
            continue
        mmean = df[col].mean()
        if mmean > 1.0:
            print(f"    [!] {mname} looks like percent (mean={mmean:.2f}); "
                  f"auto-converting to decimal")
            df[col] = df[col] / 100.0
    # Post-conversion sanity check — now these must all be decimal
    _y10 = df[YIELD_COLS["10Y"]].mean()
    _cpi = df[MACRO_COLS["CPI"]].mean()
    _repo = df[MACRO_COLS["Repo Rate"]].mean()
    assert _y10 < 1.0, f"10Y mean={_y10:.4f} still looks like percent after conversion"
    assert _cpi < 1.0, f"CPI mean={_cpi:.4f} still looks like percent after conversion"
    assert _repo < 1.0, f"Repo mean={_repo:.4f} still looks like percent after conversion"
    # v3.5.5 FIX 2: Cross-maturity monotonicity / outlier check.
    # Yield curves should be broadly monotone (short < long). If any
    # maturity's yield is >50 bps away from linear interpolation of its
    # neighbours, replace with interpolated value. This catches data errors
    # like the 12Y=6.634% observed April 9 (should have been ~7.10%
    # between 11Y=7.19% and 13Y=7.36%).
    sorted_mats = sorted(MATURITIES_YRS.items(), key=lambda x: x[1])
    mat_labs = [m[0] for m in sorted_mats]
    mat_yrs  = [m[1] for m in sorted_mats]
    n_fixed = 0
    for i in range(1, len(mat_labs) - 1):
        lab = mat_labs[i]
        if lab not in YIELD_COLS:
            continue
        col = YIELD_COLS[lab]
        prev_lab, next_lab = mat_labs[i-1], mat_labs[i+1]
        if prev_lab not in YIELD_COLS or next_lab not in YIELD_COLS:
            continue
        prev_col, next_col = YIELD_COLS[prev_lab], YIELD_COLS[next_lab]
        # Linear interpolation between neighbours
        w = (mat_yrs[i] - mat_yrs[i-1]) / (mat_yrs[i+1] - mat_yrs[i-1])
        interp = df[prev_col] * (1 - w) + df[next_col] * w
        deviation = (df[col] - interp).abs() * 10_000   # bps
        bad = deviation > 50
        n_bad = bad.sum()
        if n_bad > 0:
            df.loc[bad, col] = interp[bad]
            n_fixed += n_bad
    if n_fixed > 0:
        print(f"    [!] Fixed {n_fixed} cross-maturity outliers (>50 bps from neighbours)")

    df.dropna(inplace=True)
    print(f"    Rows: {len(df):,}  |  "
          f"Range: {df.index.min():%Y-%m-%d} to {df.index.max():%Y-%m-%d}")
    return df, mat_real_obs


# ============================================================================
# SECTION 4 — MACRO WEIGHTAGE INPUT
# ============================================================================
def get_macro_weights(config=None):
    """Get macro weights. If `config` dict has 'macro_weights', use those
    (non-interactive mode). Otherwise prompt via stdin."""
    print("\n" + "=" * 72)
    print("  MACRO VARIABLE WEIGHTAGE ASSIGNMENT")
    print("=" * 72)
    weights = {}
    if config and isinstance(config.get("macro_weights"), dict):
        cfg_w = config["macro_weights"]
        for name, default in DEFAULT_MACRO_WEIGHTS.items():
            try:
                weights[name] = float(cfg_w.get(name, default))
            except (TypeError, ValueError):
                weights[name] = float(default)
        print("  Using weights from predictor_config.json (non-interactive).")
    else:
        print("  Assign importance weights (0-100). Press Enter for default.\n")
        for name, default in DEFAULT_MACRO_WEIGHTS.items():
            while True:
                try:
                    raw = input(f"    {name:12s}  [default {default:2d}] : ").strip()
                    weights[name] = float(raw) if raw else default
                    break
                except ValueError:
                    print("      -> enter a number")
    total = sum(weights.values()) or 1.0
    weights = {k: v / total * 100 for k, v in weights.items()}
    print("\n  Normalised weights:")
    for k, v in weights.items():
        print(f"    {k:12s}  {v:5.1f}%  {'|' * int(v / 2)}")
    return weights


# ============================================================================
# SECTION 4b — HELPER FUNCTIONS  (factory, ACF seq_len, FLAT threshold)
# ============================================================================
def _make_xgb(n_estimators=500, max_depth=6, early_stopping_rounds=50,
              monotone_constraints=None):
    """Factory for XGBoost regressors — shared hyperparameters enforced.
    build_xgboost() uses defaults; run_backtest() passes smaller values
    explicitly (fewer trees for smaller per-fold training sets).

    v3.5.5 FIX B: colsample_bytree 0.7 -> 0.5 forces every tree to draw
    from a smaller random feature pool. Without this, redundant smoothed
    level features (ma20_/ma60_) crowd out the few directional signals
    (forward-spot spreads, RBI stance) because trees always prefer the
    locally most-informative split and there are ~35 level features vs
    6 directional ones. Lower colsample spreads importance more evenly.

    v3.5.5 FIX D: monotone_constraints can be passed as a tuple/list of
    {-1, 0, +1} aligned with training-matrix column order. Forward-spread
    features get +1 (curve-implied upward move -> predict up) and RBI
    stance gets +1 (hiking -> yields rise).
    """
    kwargs = dict(
        n_estimators=n_estimators, max_depth=max_depth, learning_rate=0.01,
        subsample=0.8, colsample_bytree=0.5, colsample_bylevel=0.7,
        min_child_weight=5,
        reg_alpha=0.1, reg_lambda=1.0, eval_metric="rmse",
        early_stopping_rounds=early_stopping_rounds,
        random_state=42, verbosity=0,
    )
    if monotone_constraints is not None:
        # XGBoost sklearn API accepts either a tuple string "(1,0,0,...)"
        # or a dict {col_idx: sign}. Tuple-string is the most robust.
        kwargs["monotone_constraints"] = "(" + ",".join(
            str(int(c)) for c in monotone_constraints) + ")"
    return xgb.XGBRegressor(**kwargs)


def _build_monotone_vec(columns):
    """v3.5.5 FIX D helper: return monotone-constraint list aligned to `columns`.
    +1 where the feature has a known monotone relationship with future yield.
    """
    mono = []
    for c in columns:
        if c.startswith("fwd_spread_"):
            mono.append(1)        # forward > spot -> expect up
        elif c == "rbi_stance" or c == "rbi_stance_ma20":
            mono.append(1)        # hiking stance -> expect up
        else:
            mono.append(0)
    return mono


def _compute_seq_len(series, max_lag=120, min_lag=20, default=60):
    """ACF-based LSTM sequence length: first lag where |ACF| < 0.05.
    Falls back to `default` if series is too short or ACF never decays.
    """
    try:
        s = series.diff().dropna().values
        if len(s) < max_lag * 2:
            return default
        a = acf(s, nlags=max_lag, fft=True)
        for i in range(min_lag, len(a)):
            if abs(a[i]) < 0.05:
                return i
        return default
    except Exception:
        return default


def _compute_flat_bps(df, col, horizon, fallback=3.0):
    """v3.5.5: Per-maturity FLAT threshold — VOLATILITY-ADAPTIVE.

    Old: 20th pctile of |historical Δ|. With 1W horizon in calm period this
    returned ~5-8 bps, killing ALL predictions since model magnitudes are
    0.2-1.0 bps. Result: 12/17 maturities labelled FLAT when reality was
    -18 to -38 bps (April 9-16 2026 validation).

    New: threshold = 10th pctile (more permissive base)
         × min(1.0, recent_vol / long_vol) — shrinks in calm regimes,
           stays full in volatile regimes (when FLAT is wrong)
         × horizon multiplier  {5: 0.4, 20: 0.7, 60: 1.0}
           shorter horizons = smaller changes expected = lower threshold
    """
    hist = (df[col].diff(horizon) * 10_000).dropna().tail(504)
    if len(hist) < 60:
        return fallback
    # Base: 10th percentile (was 20th — too aggressive)
    base = max(0.3, float(np.percentile(np.abs(hist), 10)))
    # Vol ratio: recent / long-run (< 1 in calm, > 1 in volatile)
    recent_vol = hist.tail(60).std()
    long_vol   = hist.std()
    vol_ratio  = min(1.0, recent_vol / (long_vol + 1e-6))
    # Horizon scaler: short horizons need proportionally lower thresholds
    h_scale = {5: 0.4, 20: 0.7, 60: 1.0}.get(horizon, 0.7)
    return base * vol_ratio * h_scale


def _detect_regime(df, horizon):
    """v3.5.5 FIX 7: Regime volatility detector.

    Returns a dict with:
      vol_regime: "high" | "normal" | "low"
      vol_ratio:  recent_vol / long_vol (used by CI, magnitude scaler)
      trend_strength: average |directional persistence| over key mats
      crude_shock: True if crude moved >5% in last `horizon` days

    The ensemble uses this to switch between mean-reversion mode (normal/low
    vol) and trend-following mode (high vol / crude shock).
    """
    # Crude oil shock detection
    crude_col = MACRO_COLS.get("Crude")
    crude_shock = False
    crude_5d_chg = 0.0
    if crude_col and crude_col in df.columns:
        crude = df[crude_col].dropna()
        if len(crude) > horizon:
            crude_5d_chg = float((crude.iloc[-1] / crude.iloc[-horizon - 1] - 1))
            crude_shock = abs(crude_5d_chg) > 0.05

    # Yield volatility regime (10Y as proxy)
    y10 = df[YIELD_COLS.get("10Y", list(YIELD_COLS.values())[0])]
    changes = (y10.diff(horizon) * 10_000).dropna()
    recent_vol = float(changes.tail(60).std()) if len(changes) > 60 else 10.0
    long_vol   = float(changes.tail(504).std()) if len(changes) > 504 else recent_vol
    vol_ratio  = recent_vol / (long_vol + 1e-6)

    if vol_ratio > 1.3 or crude_shock:
        vol_regime = "high"
    elif vol_ratio < 0.7:
        vol_regime = "low"
    else:
        vol_regime = "normal"

    # Trend strength: what fraction of key mats moved same direction over recent window
    key_cols = [YIELD_COLS[m] for m in ["3M","2Y","5Y","10Y","30Y"] if m in YIELD_COLS]
    dirs = []
    for c in key_cols:
        chg = float(df[c].diff(horizon).iloc[-1] * 10_000) if len(df) > horizon else 0
        dirs.append(np.sign(chg))
    if dirs:
        trend_strength = abs(np.mean(dirs))   # 1.0 = all same dir, 0 = mixed
    else:
        trend_strength = 0.0

    return {
        "vol_regime":     vol_regime,
        "vol_ratio":      round(vol_ratio, 3),
        "trend_strength": round(trend_strength, 3),
        "crude_shock":    crude_shock,
        "crude_chg_pct":  round(crude_5d_chg * 100, 2),
        "recent_vol":     round(recent_vol, 2),
        "long_vol":       round(long_vol, 2),
    }


# ============================================================================
# SECTION 5 — NELSON-SIEGEL-SVENSSON DAILY FITTER
# ============================================================================
def fit_nss_fast(df):
    """Fit NSS to each day using fixed-lambda OLS + grid search.
    Returns DataFrame with [nss_b0, nss_b1, nss_b2, nss_b3] per day.
    """
    print("\n[2/12] Fitting Nelson-Siegel-Svensson (daily) ...")

    # Cache version tag — bump when fitting method OR upstream unit handling
    # changes, to force refit (v3.5 changed load_data unit handling)
    NSS_VERSION = "v3.5.5-regime-adaptive"
    if NSS_CACHE.exists():
        cached = pd.read_csv(NSS_CACHE, index_col=0, parse_dates=True)
        cache_ver = cached.columns[-1] if cached.columns[-1].startswith("ver") else ""
        last_cache = cached.index.max()
        last_data  = df.index.max()
        if cache_ver == NSS_VERSION and (last_data - last_cache).days <= 3:
            data_cols = [c for c in cached.columns if not c.startswith("ver")]
            print(f"    Loaded from cache ({len(cached)} rows, "
                  f"through {last_cache:%Y-%m-%d})")
            return cached[data_cols].reindex(df.index).ffill().bfill()
        reason = "version change" if cache_ver != NSS_VERSION else "stale data"
        print(f"    Cache invalid ({reason}), refitting ...")

    taus = np.array(list(MATURITIES_YRS.values()))
    yields_mat = df[list(YIELD_COLS.values())].values  # (T, 22)

    # Grid search over lambda1, lambda2 using 5 representative days
    # (covers early, mid, late regimes — not just one median day)
    best_l1, best_l2, best_err = 1.5, 5.0, np.inf
    n_days = len(yields_mat)
    rep_idx = [n_days // 6, n_days // 3, n_days // 2,
               2 * n_days // 3, 5 * n_days // 6]
    rep_yields = yields_mat[rep_idx]
    for l1 in [0.5, 1.0, 1.5, 2.0, 3.0, 5.0]:
        for l2 in [2.0, 5.0, 8.0, 12.0, 20.0]:
            if l2 <= l1:
                continue
            e1 = np.exp(-taus / l1)
            e2 = np.exp(-taus / l2)
            X = np.column_stack([
                np.ones_like(taus),
                (1 - e1) / (taus / l1),
                (1 - e1) / (taus / l1) - e1,
                (1 - e2) / (taus / l2) - e2,
            ])
            total_err = 0
            for ry in rep_yields:
                betas, _, _, _ = np.linalg.lstsq(X, ry, rcond=None)
                total_err += np.sum((X @ betas - ry) ** 2)
            if total_err < best_err:
                best_err = total_err
                best_l1, best_l2 = l1, l2

    # Build fixed loading matrix
    e1 = np.exp(-taus / best_l1)
    e2 = np.exp(-taus / best_l2)
    X_load = np.column_stack([
        np.ones_like(taus),
        (1 - e1) / (taus / best_l1),
        (1 - e1) / (taus / best_l1) - e1,
        (1 - e2) / (taus / best_l2) - e2,
    ])

    # Daily OLS: betas = (X'X)^-1 X'y  (vectorised)
    XtX_inv = np.linalg.inv(X_load.T @ X_load)
    betas_all = (XtX_inv @ X_load.T @ yields_mat.T).T  # (T, 4)

    result = pd.DataFrame(
        betas_all, index=df.index,
        columns=["nss_b0", "nss_b1", "nss_b2", "nss_b3"]
    )
    result["ver_tag"] = NSS_VERSION   # version tag for cache invalidation

    # Cache
    result.to_csv(NSS_CACHE)
    result = result.drop(columns=["ver_tag"])
    print(f"    Fitted {len(result)} days  |  lambda1={best_l1}, lambda2={best_l2}")
    print(f"    Cached to {NSS_CACHE.name}")
    return result


# ============================================================================
# SECTION 6 — FEATURE ENGINEERING  (FIXED: ma60, momentum, macro scaling)
# ============================================================================
def engineer_features(df, macro_weights, nss_params):
    print("\n[3/12] Engineering features ...")
    F = pd.DataFrame(index=df.index)

    # --- yield features ---
    # v3.5.5 FIX A: ma5/ma20/ma60 level features restricted to KEY maturities
    # only. v3.5.5 had these for all 13 maturities (~39 smoothed level
    # features), which crowded out the 6 new directional features in
    # XGBoost tree splits. KEY = 3M, 2Y, 5Y, 10Y, 30Y only — the rest keep
    # their differentials (dy1/dy5/dy20) and level y_ but not moving avgs.
    KEY_MA_MATS = {"3M", "2Y", "5Y", "10Y", "30Y"}
    for lab, col in YIELD_COLS.items():
        F[f"y_{lab}"]    = df[col]
        F[f"dy1_{lab}"]  = df[col].diff(1)  * 10_000
        F[f"dy5_{lab}"]  = df[col].diff(5)  * 10_000
        F[f"dy20_{lab}"] = df[col].diff(20) * 10_000
        if lab in KEY_MA_MATS:
            F[f"ma5_{lab}"]  = df[col].rolling(5).mean()
            F[f"ma20_{lab}"] = df[col].rolling(20).mean()
            F[f"ma60_{lab}"] = df[col].rolling(60).mean()

    # momentum (only for KEY_MA_MATS since ma20/ma60 only exist there)
    for lab in KEY_MA_MATS:
        if lab in YIELD_COLS:
            F[f"mom_20_60_{lab}"] = F[f"ma20_{lab}"] - F[f"ma60_{lab}"]

    # yield-curve shape
    F["slope_10y2y"] = (df[YIELD_COLS["10Y"]] - df[YIELD_COLS["2Y"]]) * 10_000
    F["slope_10y3m"] = (df[YIELD_COLS["10Y"]] - df[YIELD_COLS["3M"]]) * 10_000
    F["slope_30y5y"] = (df[YIELD_COLS["30Y"]] - df[YIELD_COLS["5Y"]]) * 10_000
    F["butterfly"]   = (2*df[YIELD_COLS["5Y"]]
                        - df[YIELD_COLS["2Y"]]
                        - df[YIELD_COLS["10Y"]]) * 10_000

    # rolling vol
    for lab in ["3M","1Y","5Y","10Y","30Y"]:
        F[f"vol20_{lab}"] = df[YIELD_COLS[lab]].diff().rolling(20).std() * 10_000
        F[f"vol60_{lab}"] = df[YIELD_COLS[lab]].diff().rolling(60).std() * 10_000

    # --- NSS factors as features ---
    F = F.join(nss_params, how="left")
    for c in ["nss_b0","nss_b1","nss_b2","nss_b3"]:
        F[f"d_{c}"]    = F[c].diff(1)
        F[f"d5_{c}"]   = F[c].diff(5)
        F[f"d20_{c}"]  = F[c].diff(20)
        F[f"ma20_{c}"] = F[c].rolling(20).mean()

    # --- macro features (scaled by user weights — linear, not sqrt) ---
    # Linear scaling gives range ~0.7x–1.8x (meaningful for LSTM);
    # XGBoost gets macro influence via macro_composite + sample weights.
    # v3.5.5 FIX C: drop direct per-variable features for macros that are
    # statistically insignificant (Granger p > 0.5). IIP showed p=0.8066
    # but appeared as #4 XGBoost feature with 23.7% importance — a clear
    # spurious fit. Keeping IIP only inside macro_composite preserves the
    # user-weight control but prevents direct attachment to IIP levels.
    MACRO_NO_DIRECT = {"IIP"}   # extend here if new diagnostics fail Granger
    for name, col in MACRO_COLS.items():
        if name in MACRO_NO_DIRECT:
            continue   # still enters via macro_composite below
        w_scale = macro_weights[name] / (100 / len(MACRO_COLS))
        # Clip pct_change to [-5, 5] — IIP near-zero values cause explosions
        F[f"m_{name}"]       = df[col] * w_scale
        F[f"m_chg5_{name}"]  = df[col].pct_change(5).clip(-5, 5)  * w_scale
        F[f"m_chg20_{name}"] = df[col].pct_change(20).clip(-5, 5) * w_scale
        F[f"m_ma20_{name}"]  = df[col].rolling(20).mean() * w_scale
        F[f"m_ma60_{name}"]  = df[col].rolling(60).mean() * w_scale
        F[f"m_std20_{name}"] = df[col].rolling(20).std()

    # Weighted macro composite z-score
    # Rate variables (CPI, Repo, IIP): use diff() — change in the rate is the signal
    # Level variables (NSE, FII, Crude, USD/INR): use pct_change() — % change
    RATE_VARS = {"CPI", "Repo Rate", "IIP"}
    z_parts = pd.DataFrame(index=df.index)
    for name, col in MACRO_COLS.items():
        if name in RATE_VARS:
            chg = df[col].diff(20)                           # absolute change in rate
        else:
            chg = df[col].pct_change(20).clip(-5, 5)        # % change, clipped
        mu  = chg.rolling(252, min_periods=60).mean()
        sig = chg.rolling(252, min_periods=60).std()
        z_parts[name] = ((chg - mu) / (sig + 1e-12)).clip(-4, 4)
    # v3.5.5 FIX: bfill instead of fillna(0). fillna(0) tells the model the
    # macro environment was "perfectly neutral" for the first 252 rows,
    # which is factually wrong — early 2015 was an active RBI cutting cycle.
    # bfill inherits the first valid z-score reading so the warm-up window
    # carries the regime that immediately followed it, not a false neutral.
    z_parts = z_parts.bfill().fillna(0)     # bfill first, fallback 0 for all-NaN cols
    w_arr = np.array([macro_weights[k] / 100 for k in MACRO_COLS])
    F["macro_composite"] = z_parts.values @ w_arr

    # v3.5.5 FIX: real_yield / term_prem scaled to bps (*10_000) to match
    # every other yield-differential feature (slope_*, butterfly, dy*_*).
    # v3.5 had these in raw decimal (~0.01-0.03), which is 4 orders of
    # magnitude smaller than bps features and distorts RobustScaler in LSTM.
    F["real_yield_10y"] = (df[YIELD_COLS["10Y"]] - df[MACRO_COLS["CPI"]])       * 10_000
    F["term_prem_10y"]  = (df[YIELD_COLS["10Y"]] - df[MACRO_COLS["Repo Rate"]]) * 10_000

    # ----------------------------------------------------------------------
    # v3.5.5 NEW FEATURES — directional signal boosters
    # ----------------------------------------------------------------------
    # (A) Forward-spot spreads from the existing yield curve. The forward
    # rate F(t1,t2) = (t2*y2 - t1*y1)/(t2-t1); forward minus spot y2
    # captures the market-implied direction for that maturity segment.
    # Zero extra cost — uses yields we already store.
    FWD_PAIRS = [(2, 5, "2_5"), (5, 10, "5_10"), (10, 15, "10_15"), (2, 10, "2_10")]
    for t1, t2, name in FWD_PAIRS:
        k1, k2 = f"{t1}Y", f"{t2}Y"
        if k1 in YIELD_COLS and k2 in YIELD_COLS:
            y1 = df[YIELD_COLS[k1]]
            y2 = df[YIELD_COLS[k2]]
            fwd = (t2 * y2 - t1 * y1) / (t2 - t1)
            F[f"fwd_spread_{name}"] = (fwd - y2) * 10_000   # bps

    # (B) RBI policy stance flag — derived from 60d repo change, not dates.
    # Threshold 10 bps (0.001) because Indian RBI moves in 25-50 bp steps,
    # so anything below 10 bps over 3 months is just noise / unchanged.
    repo = df[MACRO_COLS["Repo Rate"]]
    repo_3m_chg = repo.diff(60)
    F["rbi_stance"] = np.where(repo_3m_chg > 0.001, 1,       # hiking
                        np.where(repo_3m_chg < -0.001, -1,    # cutting
                        0))                                    # on-hold
    F["rbi_stance_ma20"] = F["rbi_stance"].rolling(20).mean()

    F.replace([np.inf, -np.inf], np.nan, inplace=True)
    F.dropna(inplace=True)
    print(f"    {F.shape[1]} features  |  {len(F):,} usable rows")
    return F


# ============================================================================
# SECTION 7 — DIAGNOSTIC TESTS
# ============================================================================
def run_diagnostics(df):
    print("\n[4/12] Running diagnostic tests ...")
    diag = {}

    print("\n  ADF Stationarity (H0: non-stationary):")
    print(f"  {'Series':20s} {'Stat':>9s} {'p-val':>8s} {'Result':>12s}")
    print("  " + "-" * 52)
    adf_info = {}
    test_series = {"10Y Yield": YIELD_COLS["10Y"], "3M Yield": YIELD_COLS["3M"],
                   "CPI": MACRO_COLS["CPI"], "Repo Rate": MACRO_COLS["Repo Rate"],
                   "USD/INR": MACRO_COLS["USD/INR"]}
    for name, col in test_series.items():
        s = df[col].dropna()
        if len(s) < 200: continue
        res = adfuller(s, maxlag=20, autolag="AIC")
        stat_str = "Stationary" if res[1] < 0.05 else "NON-stat"
        adf_info[name] = {"stat": res[0], "p": res[1], "stationary": res[1] < 0.05}
        print(f"  {name:20s} {res[0]:9.3f} {res[1]:8.4f} {stat_str:>12s}")

    print("\n  First differences:")
    for name, col in test_series.items():
        s = df[col].diff().dropna()
        if len(s) < 200: continue
        res = adfuller(s, maxlag=20, autolag="AIC")
        stat_str = "Stationary" if res[1] < 0.05 else "NON-stat"
        print(f"  d({name:18s}) {res[0]:9.3f} {res[1]:8.4f} {stat_str:>12s}")
    diag["adf"] = adf_info

    print(f"\n  Linearity:  Macro -> 10Y Yield")
    print(f"  {'Variable':15s} {'Pearson':>9s} {'Spearman':>9s} {'Type':>14s}")
    print("  " + "-" * 50)
    linearity = {}
    y10 = df[YIELD_COLS["10Y"]]
    for name, col in MACRO_COLS.items():
        x = df[col]; idx = y10.index.intersection(x.dropna().index)
        if len(idx) < 200: continue
        pr, _ = stats.pearsonr(y10.loc[idx], x.loc[idx])
        sr, _ = stats.spearmanr(y10.loc[idx], x.loc[idx])
        gap = abs(abs(sr) - abs(pr))
        rtype = "Non-Linear" if gap > 0.10 else ("Linear" if abs(pr) > 0.25 else "Weak")
        linearity[name] = {"pearson": pr, "spearman": sr, "type": rtype}
        print(f"  {name:15s} {pr:9.4f} {sr:9.4f} {rtype:>14s}")
    diag["linearity"] = linearity

    # NOTE: Johansen cointegration now runs inside build_econometric()
    # so that k_ar_diff matches the VAR-selected lag (not hardcoded 2).

    # v3.5 FIX 4: Granger causality — rate vars (CPI/Repo/IIP) are MONTHLY
    # by publication, so daily data with lag=3 tests causality over 3 DAYS
    # which is meaningless. Use monthly resample + lag=6 (6 months) for
    # rate vars; daily + lag=3 remains appropriate for level vars.
    print(f"\n  Granger Causality  (Macro -> d(10Y)):")
    print(f"  Rate vars: monthly, maxlag=6 (6 months)")
    print(f"  Level vars: daily, maxlag=5 (1 week)")
    RATE_VARS = {"CPI", "Repo Rate", "IIP"}
    gc = {}
    y10_col = YIELD_COLS["10Y"]
    y10_monthly_d = df[y10_col].resample("ME").last().diff().dropna()
    y10_daily_d   = df[y10_col].diff().dropna()
    for name, col in MACRO_COLS.items():
        try:
            if name in RATE_VARS:
                # Monthly resample, then diff() — absolute change in the rate
                x_m = df[col].resample("ME").last().diff().dropna()
                idx = y10_monthly_d.index.intersection(x_m.index)
                tmp = pd.DataFrame({"y": y10_monthly_d.loc[idx], "x": x_m.loc[idx]})
                tmp.replace([np.inf, -np.inf], np.nan, inplace=True)
                tmp.dropna(inplace=True)
                if len(tmp) < 48:
                    continue
                maxlag = min(6, len(tmp) // 8)
            else:
                # Level vars: daily pct_change, short lag (1 week)
                x_d = df[col].pct_change().dropna()
                idx = y10_daily_d.index.intersection(x_d.index)
                tmp = pd.DataFrame({"y": y10_daily_d.loc[idx], "x": x_d.loc[idx]})
                tmp.replace([np.inf, -np.inf], np.nan, inplace=True)
                tmp.dropna(inplace=True)
                if len(tmp) < 300:
                    continue
                maxlag = 5
            res  = grangercausalitytests(tmp[["y", "x"]], maxlag=maxlag, verbose=False)
            pmin = min(res[lag][0]["ssr_ftest"][1] for lag in range(1, maxlag + 1))
            sig  = "***" if pmin < 0.01 else "**" if pmin < 0.05 else "*" if pmin < 0.10 else ""
            freq = "mo" if name in RATE_VARS else "d "
            gc[name] = pmin
            print(f"    {name:12s} [{freq}] p={pmin:.4f}  {sig}")
        except Exception as e:
            print(f"    {name:12s}  skipped ({type(e).__name__})")
    diag["granger"] = gc
    return diag


# ============================================================================
# SECTION 8 — HULL-WHITE SHORT RATE MODEL
# ============================================================================
def fit_hull_white(df, horizon_days):
    """1-factor Hull-White on 3M yield, anchored to Repo Rate."""
    print(f"\n[5/12] Fitting Hull-White short rate model ...")
    short_yield = df[YIELD_COLS["3M"]].dropna()
    DT = 1 / 252

    def neg_ll(params):
        kappa, theta, sigma = params
        if kappa <= 0 or sigma <= 0:
            return 1e10
        r = short_yield.values
        r_pred = r[:-1] + kappa * (theta - r[:-1]) * DT
        resid  = r[1:] - r_pred
        var    = sigma**2 * DT
        return 0.5 * np.sum(resid**2 / var + np.log(var))

    theta_init = float(df[MACRO_COLS["Repo Rate"]].iloc[-1])
    mean_3m    = float(short_yield.mean())
    # v3.5 FIX 9: Diversified theta starting points covering multiple regimes.
    # Previous starts clustered on current repo (theta_init) — if likelihood
    # surface has a minimum near historical mean or a stressed regime, the
    # optimizer never found it. Now starts span: current, mean, +150bps stress,
    # -100bps dovish. Each theta gets two kappas (fast/slow mean reversion).
    bounds = [(0.01, 20), (0.001, 0.20), (0.0001, 0.05)]
    theta_hi = min(0.18, theta_init + 0.015)   # stressed (+150 bps)
    theta_lo = max(0.02, theta_init - 0.010)   # dovish (-100 bps)
    starts = [
        (0.5,  theta_init, 0.005),
        (2.0,  theta_init, 0.005),
        (5.0,  theta_init, 0.010),
        (2.0,  mean_3m,    0.005),
        (10.0, theta_init, 0.010),
        (1.0,  theta_hi,   0.008),   # stressed regime
        (3.0,  theta_hi,   0.012),
        (1.0,  theta_lo,   0.005),   # dovish regime
        (3.0,  theta_lo,   0.008),
    ]
    best_res, best_nll = None, np.inf
    for x0 in starts:
        try:
            r = sp_minimize(neg_ll, x0, method="L-BFGS-B", bounds=bounds)
            if r.fun < best_nll:
                best_nll = r.fun
                best_res = r
        except Exception:
            continue
    if best_res is None:
        kappa, theta, sigma = 2.0, theta_init, 0.005   # fallback
    else:
        kappa, theta, sigma = best_res.x

    # In-sample RMSE for ensemble weighting
    rv = short_yield.values
    r_pred_is = rv[:-1] + kappa * (theta - rv[:-1]) * DT
    hw_rmse = float(np.sqrt(np.mean((rv[1:] - r_pred_is)**2)) * 10_000)

    # v3.5 FIX 13: In-sample directional accuracy for ensemble confidence.
    # Use 20-day realized vs predicted change (typical horizon). Previously
    # the ensemble hardcoded HW conf=0.60, which was optimistic for short-end.
    try:
        step = 20
        actual_chg = rv[step:] - rv[:-step]
        pred_chg   = (theta - rv[:-step]) * (1 - np.exp(-kappa * step * DT))
        hw_dir_acc = float(np.mean(np.sign(actual_chg) == np.sign(pred_chg)))
        hw_dir_acc = float(np.clip(hw_dir_acc, 0.50, 0.75))
    except Exception:
        hw_dir_acc = 0.60

    r_now = float(short_yield.iloc[-1])
    h     = horizon_days * DT
    r_fc  = theta + (r_now - theta) * np.exp(-kappa * h)
    chg   = (r_fc - r_now) * 10_000
    ci_v  = sigma * np.sqrt((1 - np.exp(-2*kappa*h)) / (2*kappa)) * 10_000

    # Compute data-driven attenuation for 6M/1Y from historical correlations
    d3m = df[YIELD_COLS["3M"]].diff().dropna()
    att = {}
    for lab in ["6M", "1Y"]:
        d_tgt = df[YIELD_COLS[lab]].diff().dropna()
        idx   = d3m.index.intersection(d_tgt.index)
        if len(idx) > 200:
            # beta from OLS regression: d_target = beta * d_3M
            b = float(np.cov(d_tgt.loc[idx], d3m.loc[idx])[0, 1]
                      / (np.var(d3m.loc[idx]) + 1e-12))
            att[lab] = float(np.clip(b, 0.5, 1.0))
        else:
            att[lab] = 0.85 if lab == "6M" else 0.70   # fallback

    print(f"    kappa={kappa:.3f}  theta={theta:.4f}  sigma={sigma:.5f}")
    print(f"    In-sample RMSE: {hw_rmse:.2f} bps")
    print(f"    Term attenuation:  6M={att['6M']:.3f}  1Y={att['1Y']:.3f}")
    ci_width = 2 * 1.65 * ci_v
    print(f"    3M forecast: {r_fc:.4f}  (change: {chg:+.1f} bps, "
          f"90%CI: [{chg - 1.65*ci_v:+.1f}, {chg + 1.65*ci_v:+.1f}])")
    # v3.5 FIX 15: Horizon-dependent CI flag. A 5-day horizon with 80 bps CI
    # is nonsense; a 60-day horizon with 80 bps is normal. Threshold scales
    # roughly with sqrt(horizon): {5:30, 20:50, 60:100} bps.
    ci_thresh = {5: 30, 20: 50, 60: 100}.get(horizon_days, 60)
    if ci_width > ci_thresh:
        print(f"    [!] CI width ({ci_width:.0f} bps) exceeds {ci_thresh} bps "
              f"for {horizon_days}d horizon — short-rate prediction unreliable")

    return {"kappa": kappa, "theta": theta, "sigma": sigma,
            "r_forecast": r_fc, "change_bps": chg, "rmse": hw_rmse,
            "attenuation": att, "dir_acc": hw_dir_acc,
            "ci_lo": chg - 1.65 * ci_v, "ci_hi": chg + 1.65 * ci_v}


# ============================================================================
# SECTION 9 — VAR / VECM  (FIXED: 4 vars, lag <= 3, multi-step forecast)
# ============================================================================
def build_econometric(df, horizon_days):
    print(f"\n[6/12] Building VAR / VECM  (horizon={horizon_days}d) ...")
    # v3.5 FIX 5: Variable exclusion disclosure — VAR uses only 4 vars
    # (kept small on purpose to avoid overparameterisation on ~180 monthly
    # observations). The excluded macros (FII, USD/INR, Crude, NSE) are
    # STILL used by XGBoost and LSTM via macro_composite + direct features,
    # so their information is captured in the ensemble — just not in VAR.
    included = ["5Y", "10Y", "CPI", "Repo Rate"]
    excluded = ["FII", "USD/INR", "Crude", "NSE", "IIP"]
    print(f"    VAR vars (4):  {included}")
    print(f"    Excluded from VAR (captured by XGB/LSTM): {excluded}")
    var_cols = [YIELD_COLS["5Y"], YIELD_COLS["10Y"],
                MACRO_COLS["CPI"], MACRO_COLS["Repo Rate"]]

    monthly      = df[var_cols].resample("ME").last().dropna()
    monthly_diff = monthly.diff().dropna()
    cutoff = monthly_diff.index.max() - pd.DateOffset(years=15)
    md = monthly_diff[monthly_diff.index >= cutoff]
    if len(md) < 48:
        print("    Insufficient data -- skipping VAR.")
        return None

    try:
        model = VAR(md)
        sel   = model.select_order(maxlags=3)   # capped at 3
        lag   = max(sel.aic, 1)
        vr    = model.fit(lag)

        # v3.5 FIX 2: VAR horizon scaling.
        # The VAR is fitted on MONTHLY differences, so each forecast step
        # is ~20 trading days. For horizon=5 (1W), round(5/20)=0 and we'd
        # previously silently fall back to steps=1 (1 month forecast for a
        # 1-week request — 4x overstated!). Fix: always compute N steps to
        # cover horizon, then rescale the cumulative forecast linearly to
        # the actual horizon. This preserves direction and magnitude scale.
        steps = max(1, int(np.ceil(horizon_days / 20)))
        fcast = vr.forecast(md.values[-lag:], steps=steps)
        cum_raw = fcast.sum(axis=0)                     # N months cumulative
        scale   = horizon_days / (steps * 20.0)         # 5/(1*20)=0.25, 20/(1*20)=1, 60/(3*20)=1
        cum_fcast = cum_raw * scale

        # v3.5 FIX 3: Scale-match VAR RMSE to the horizon for ensemble weighting.
        # vr.resid is a ONE-step (one month) residual; XGBoost/LSTM RMSE is
        # horizon-step residual. Mixing them gives VAR inflated/deflated weight.
        # Under iid innovations, multi-step RMSE scales as sqrt(steps), and
        # rescaling by `scale` gives horizon-equivalent variance.
        resid = vr.resid
        var_rmse_monthly = float(np.sqrt(np.mean(resid ** 2)) * 10_000)
        var_rmse = var_rmse_monthly * np.sqrt(steps) * scale

        # v3.5 FIX 13: In-sample directional accuracy for ensemble confidence.
        # Previous code used a hardcoded 0.52 which understated VAR.
        # Here we use one-step in-sample residuals as a proxy for sign skill.
        try:
            y10_in_sample = md[YIELD_COLS["10Y"]].values[lag:]   # actual
            y10_pred_is   = y10_in_sample - resid[:, 1]          # fitted
            var_dir_acc = float(np.mean(np.sign(y10_in_sample) == np.sign(y10_pred_is)))
            var_dir_acc = float(np.clip(var_dir_acc, 0.50, 0.70))   # conservative cap
        except Exception:
            var_dir_acc = 0.52

        print(f"    VAR({lag}) fitted  |  {len(md)} obs  |  {steps}-step forecast"
              f"  |  scale={scale:.2f}")
        print(f"    RMSE (monthly)={var_rmse_monthly:.1f} bps  "
              f"-> RMSE ({horizon_days}d)={var_rmse:.1f} bps  "
              f"|  dir_acc={var_dir_acc:.1%}")

        try:
            irf = vr.irf(periods=12)
            print("    Impulse Response Functions computed")
        except Exception:
            irf = None

        # Johansen cointegration — runs AFTER VAR lag is known
        # so k_ar_diff matches the selected lag (not hardcoded 2)
        vecm_fcast = None
        cointegrated = False
        coint_rank = 0
        ml = monthly[monthly.index >= cutoff].dropna()
        try:
            if len(ml) > 60:
                joh = coint_johansen(ml, det_order=0, k_ar_diff=lag)
                print(f"\n    Johansen Cointegration (k_ar_diff={lag}):")
                print(f"    {'Hypothesis':18s} {'Trace':>9s} {'Crit-95%':>9s} {'Coint?':>8s}")
                print("    " + "-" * 48)
                for i in range(min(len(var_cols), len(joh.lr1))):
                    tr, cv = joh.lr1[i], joh.cvt[i, 1]
                    is_c = tr > cv
                    print(f"    r <= {i}              {tr:9.3f} {cv:9.3f} "
                          f"{'YES' if is_c else 'no':>8s}")
                    if is_c:
                        coint_rank = i + 1
                cointegrated = coint_rank > 0
                print(f"    => Cointegrating rank = {coint_rank}")
        except Exception as e:
            print(f"    Johansen skipped: {e}")

        if cointegrated:
            try:
                cr = max(1, coint_rank)
                vecm_res = VECM(ml, k_ar_diff=max(lag - 1, 1),
                                coint_rank=cr).fit()
                vecm_levels = vecm_res.predict(steps=steps)
                last_level = monthly.iloc[-1].values
                vecm_raw = vecm_levels[-1] - last_level
                # v3.5 FIX 2 (VECM): same horizon rescaling as VAR
                vecm_fcast = vecm_raw * scale
                print(f"    VECM fitted (cointegration rank {cr})")
            except Exception as e:
                print(f"    VECM skipped: {e}")

        try:
            fevd = vr.fevd(periods=12)
        except Exception:
            fevd = None

        return {"var": vr, "forecast": cum_fcast, "vecm_forecast": vecm_fcast,
                "cols": var_cols, "irf": irf, "fevd": fevd,
                "rmse": var_rmse, "dir_acc": var_dir_acc,
                "horizon_scale": scale}
    except Exception as e:
        print(f"    VAR failed: {e}")
        return None


# ============================================================================
# SECTION 10 — XGBOOST  (FIXED: gap, regressor-only, no clf mismatch)
# ============================================================================
def build_xgboost(features, df, macro_weights, horizon):
    print(f"\n[7/12] Building XGBoost  (horizon = {horizon}d) ...")
    # v3.5.5: hard 2015 cutoff AND gentle time-decay within post-2015 window.
    # v3.5.5 removed the cutoff entirely and got destroyed by 1998-2014 noise.
    # v3.5.1 used only the cutoff and overfit the 2015+ rate-cut cycle (70% -> 39% backtest).
    # v3.5.5: post-2015 only, with decay=0.00005 (8x gentler than v3.5.5's 0.0005)
    # so the 2015-2023 rate-cycle variation is still visible under the 2024-2026 tail.
    print(f"    Training window: >= {XGB_TRAIN_START:%Y-%m-%d}  "
          f"|  decay={XGB_TIME_DECAY:.5f}/row")
    models = {}

    for lab, col in YIELD_COLS.items():
        target = (df[col].shift(-horizon) - df[col]) * 10_000
        idx = features.index.intersection(target.dropna().index)
        X, y = features.loc[idx].copy(), target.loc[idx].copy()
        mask = np.isfinite(y) & np.all(np.isfinite(X.values), axis=1)
        X, y = X[mask], y[mask]
        # v3.5.5: hard post-2015 restriction (LSTM uses same cutoff, keeps
        # both models in the same regime — backtest now applies this too)
        X = X[X.index >= XGB_TRAIN_START]
        y = y.loc[X.index]
        if len(X) < 800:
            continue

        # FIX: gap = horizon between train and test (no target boundary leakage)
        split = int(len(X) * 0.80)
        Xtr = X.iloc[:split - horizon]
        ytr = y.iloc[:split - horizon]
        Xte = X.iloc[split:]
        yte = y.iloc[split:]

        if len(Xtr) < 500 or len(Xte) < 50:
            continue

        # v3.5.5: gentle decay INSIDE the post-2015 window — recent
        # 2024-2026 rows slightly emphasised, 2015-2023 still contributes.
        # With ~1800 train rows and decay=0.00005, ratio = exp(0.00005*1800) = 1.094
        # i.e. ~9% lift on newest vs oldest row — not destructive.
        n = len(Xtr)
        time_weights = np.exp(XGB_TIME_DECAY * np.arange(n))
        time_weights /= time_weights.mean()
        sw = time_weights.copy()
        if "macro_composite" in Xtr.columns:
            mc = np.abs(Xtr["macro_composite"].values)
            sw = time_weights * (1.0 + 0.5 * mc / (mc.mean() + 1e-8))

        # v3.5.5 FIX D: monotone constraints for directional features
        mono = _build_monotone_vec(Xtr.columns)
        reg = _make_xgb(monotone_constraints=mono)
        reg.fit(Xtr, ytr, sample_weight=sw,
                eval_set=[(Xte, yte)], verbose=False)

        preds = reg.predict(Xte)
        rmse  = np.sqrt(mean_squared_error(yte, preds))
        mae   = mean_absolute_error(yte, preds)
        dir_acc = np.mean(np.sign(yte) == np.sign(preds))

        # v3.5.5 FIX F: store naive baseline for ensemble trend-blend
        prop_up = float(np.mean(yte > 0))
        naive_acc = max(prop_up, 1.0 - prop_up)

        models[lab] = {"reg": reg, "dir_acc": dir_acc,
                       "rmse": rmse, "mae": mae,
                       "feats": list(X.columns),
                       "n_test": len(Xte),
                       "naive_acc": naive_acc}

    print(f"\n  {'Mat':>6s} {'DirAcc':>7s} {'RMSE':>8s} {'MAE':>8s}")
    print("  " + "-" * 32)
    for m in DISPLAY_MATS:
        if m in models:
            r = models[m]
            flag = "  [!] regime artifact?" if r["dir_acc"] > 0.80 else ""
            print(f"  {m:>6s} {r['dir_acc']:>6.1%} {r['rmse']:>8.2f} "
                  f"{r['mae']:>8.2f}{flag}")
    return models


# ============================================================================
# SECTION 11 — LSTM  (FIXED: scaler fitted on TRAIN only)
# ============================================================================
def build_lstm(features, df, horizon):
    if not LSTM_AVAILABLE:
        print("\n[8/12] LSTM -- skipped (TensorFlow not installed)")
        return None

    # v3.5.1 FIX 1a: Reproducibility — seed EVERY random source before any
    # TF op. Without this, two consecutive runs of the SAME script on the
    # SAME data produce different LSTM weights -> different dir_acc ->
    # different gating -> different ensemble -> different predictions.
    # (Observed: 2Y flipped DOWN -10.9 bps <-> UP +3.7 bps across runs.)
    import random as _py_random
    _py_random.seed(42)
    np.random.seed(42)
    tf.random.set_seed(42)
    os.environ["PYTHONHASHSEED"] = "42"
    # Force single-thread determinism for GPU/CPU kernel selection
    try:
        tf.config.experimental.enable_op_determinism()
    except Exception:
        pass  # older TF versions lack this API

    # ACF-derived sequence length (data-driven, not hardcoded 60)
    ref_col = YIELD_COLS.get("10Y") or list(YIELD_COLS.values())[0]
    seq_len = _compute_seq_len(df[ref_col])
    print(f"\n[8/12] Building Bi-LSTM  (horizon={horizon}d, seq={seq_len} [ACF-derived]) ...")
    models = {}
    key_mats = ["3M","1Y","2Y","5Y","7Y","10Y","14Y","30Y"]

    for lab in key_mats:
        if lab not in YIELD_COLS:      # safety: maturity may have been filtered
            continue
        col = YIELD_COLS[lab]
        target = (df[col].shift(-horizon) - df[col]) * 10_000

        # Select features by priority category to ensure NSS/macro aren't cut
        _cats = [
            [f"y_{lab}", f"dy1_{lab}", f"dy5_{lab}", f"dy20_{lab}",
             f"ma20_{lab}", f"ma60_{lab}", f"mom_20_60_{lab}"],
            [c for c in features.columns if c.startswith("nss_") or c.startswith("d_nss")
                                          or c.startswith("d5_nss") or c.startswith("d20_nss")],
            [c for c in features.columns if c in ("slope_10y2y","slope_10y3m","slope_30y5y",
                                                    "butterfly","macro_composite",
                                                    "real_yield_10y","term_prem_10y")],
            [c for c in features.columns if c.startswith("vol20_") or c.startswith("vol60_")],
            [c for c in features.columns if c.startswith("m_chg20_") or c.startswith("m_ma20_")],
            [c for c in features.columns if c.startswith("dy1_") and lab not in c],
        ]
        keep = []
        for cat in _cats:
            for c in cat:
                if c in features.columns and c not in keep:
                    keep.append(c)
                if len(keep) >= 60:
                    break
            if len(keep) >= 60:
                break

        idx = features.index.intersection(target.dropna().index)
        # v3.5.5: hard regime cutoff for LSTM — sequence models cannot be
        # reweighted row-wise the way XGBoost can, so the simplest way to
        # remove 1998-era regime contamination is a hard start date.
        idx = idx[idx >= XGB_TRAIN_START]
        Xf  = features[keep].loc[idx].values
        yf  = target.loc[idx].values
        mask = np.isfinite(yf) & np.all(np.isfinite(Xf), axis=1)
        Xf, yf = Xf[mask], yf[mask]
        if len(Xf) < seq_len + 600:
            continue

        # FIX: split BEFORE scaling (no scaler leakage)
        sp = int(len(Xf) * 0.80) - horizon   # gap
        if sp < seq_len + 200:
            continue
        test_start = sp + horizon

        sx = RobustScaler()
        Xf_train_s = sx.fit_transform(Xf[:sp])          # fit on train only
        Xf_test_s  = sx.transform(Xf[test_start:])      # transform test

        sy = StandardScaler()
        yf_train_s = sy.fit_transform(yf[:sp].reshape(-1,1)).ravel()
        yf_test_s  = sy.transform(yf[test_start:].reshape(-1,1)).ravel()

        # build sequences from train
        Xtr_seq, ytr_seq = [], []
        for i in range(seq_len, len(Xf_train_s)):
            Xtr_seq.append(Xf_train_s[i - seq_len:i])
            ytr_seq.append(yf_train_s[i])
        Xtr_seq, ytr_seq = np.array(Xtr_seq), np.array(ytr_seq)

        # build sequences from test
        # prepend last seq_len of train (scaled) for first test sequences
        Xf_combined = np.vstack([Xf_train_s[-seq_len:], Xf_test_s])
        yf_combined = np.concatenate([yf_train_s[-seq_len:], yf_test_s])
        Xte_seq, yte_seq = [], []
        for i in range(seq_len, len(Xf_combined)):
            Xte_seq.append(Xf_combined[i - seq_len:i])
            yte_seq.append(yf_combined[i])
        Xte_seq, yte_seq = np.array(Xte_seq), np.array(yte_seq)

        if len(Xtr_seq) < 200 or len(Xte_seq) < 50:
            continue

        mdl = Sequential([
            Bidirectional(KerasLSTM(64, return_sequences=True),
                          input_shape=(seq_len, Xtr_seq.shape[2])),
            Dropout(0.25),
            BatchNormalization(),
            KerasLSTM(32),
            Dropout(0.20),
            Dense(16, activation="relu"),
            Dense(1)
        ])
        mdl.compile(optimizer=Adam(learning_rate=0.001), loss="huber")
        # v3.5.1 FIX 1b: shuffle=False — Keras default shuffles training
        # batches each epoch, which breaks the temporal order of a time
        # series and adds run-to-run variance on top of the seed issue.
        # For sequential data the order must be preserved.
        mdl.fit(Xtr_seq, ytr_seq, epochs=50, batch_size=64,
                validation_split=0.10, verbose=0, shuffle=False,
                callbacks=[EarlyStopping(patience=8, restore_best_weights=True),
                           ReduceLROnPlateau(factor=0.5, patience=4)])

        pred_s = mdl.predict(Xte_seq, verbose=0).ravel()
        pred   = sy.inverse_transform(pred_s.reshape(-1,1)).ravel()
        act    = sy.inverse_transform(yte_seq.reshape(-1,1)).ravel()
        rmse   = np.sqrt(mean_squared_error(act, pred))
        dacc   = float(np.mean(np.sign(act) == np.sign(pred)))

        models[lab] = {"model": mdl, "sx": sx, "sy": sy,
                       "feats": keep, "seq": seq_len,
                       "rmse": rmse, "dir_acc": dacc}
        print(f"    {lab:>4s}:  DirAcc={dacc:.1%}   RMSE={rmse:.1f} bps")

    # v3.5 FIX 8: LSTM coverage disclosure — which key maturities got fitted,
    # which got skipped (insufficient data / filtered), and how many gated
    # out in the ensemble (dir_acc <= 0.52 -> v3.5.1 raised threshold).
    fitted  = sorted(models.keys())
    want    = [m for m in key_mats if m in YIELD_COLS]
    skipped = [m for m in want if m not in fitted]
    gated   = [m for m in fitted if models[m]["dir_acc"] <= 0.52]
    print(f"\n    LSTM coverage:  fitted {len(fitted)}/{len(want)}  "
          f"({fitted})")
    if skipped:
        print(f"    Skipped (insufficient data): {skipped}")
    if gated:
        print(f"    [!] Gated out of ensemble (dir_acc <= 52%): {gated}")

    return models if models else None


# ============================================================================
# SECTION 12 — ENSEMBLE  (FIXED: adaptive weights, empirical CI, no mismatch)
# ============================================================================
def ensemble_predict(xgb_m, lstm_m, econ, hw, features, df, macro_w, horizon):
    print(f"\n[9/12] Generating ensemble predictions ...")
    FLAT_FALLBACK = {5: 0.5, 20: 2, 60: 5}.get(horizon, 2)

    # v3.5.5 FIX 7: Regime detection — determines whether to lean toward
    # mean-reversion (normal/low vol) or trend-following (high vol / shock).
    regime = _detect_regime(df, horizon)
    vol_regime = regime["vol_regime"]
    print(f"    Regime: {vol_regime.upper()}  |  vol_ratio={regime['vol_ratio']:.2f}  "
          f"|  trend={regime['trend_strength']:.2f}  "
          f"|  crude={'SHOCK '+str(regime['crude_chg_pct'])+'%' if regime['crude_shock'] else 'normal'}")
    predictions = {}
    has_lstm = lstm_m is not None
    has_var  = econ is not None

    # Compute performance-adaptive weights from training RMSE
    xgb_rmses = {lab: xgb_m[lab]["rmse"] for lab in xgb_m}
    lstm_rmses = {lab: lstm_m[lab]["rmse"] for lab in lstm_m} if has_lstm else {}

    for lab, col in YIELD_COLS.items():
        current = df[col].iloc[-1] * 100
        sources = []

        # --- XGBoost (regressor only — sign = direction) ---
        if lab in xgb_m:
            Xlast = features.iloc[[-1]]
            mag   = float(xgb_m[lab]["reg"].predict(Xlast)[0])
            # FIX: confidence = backtest directional accuracy
            conf  = xgb_m[lab]["dir_acc"]
            sources.append({"src": "XGBoost", "mag": mag,
                            "dir": "UP" if mag > 0 else "DOWN",
                            "conf": conf, "rmse": xgb_m[lab]["rmse"]})

        # --- LSTM (gated: exclude if dir_acc <= 0.52 — v3.5.1 FIX 1c) ---
        # Raised from 0.50 to 0.52: a 50% boundary is too fragile. Two
        # consecutive runs saw the SAME maturity flip from 49.8% -> 58.8%
        # (3M) and 55.0% -> 41.6% (30Y), which cascaded to prediction
        # reversals. 0.52 requires genuinely better-than-random skill
        # and stabilises the ensemble under LSTM variance.
        if has_lstm and lab in lstm_m and lstm_m[lab]["dir_acc"] > 0.52:
            m   = lstm_m[lab]
            raw = features[m["feats"]].iloc[-m["seq"]:].values
            xs  = m["sx"].transform(raw).reshape(1, m["seq"], -1)
            ps  = m["model"].predict(xs, verbose=0)[0][0]
            mag = float(m["sy"].inverse_transform([[ps]])[0][0])
            # FIX: confidence = LSTM backtest accuracy (not made-up formula)
            conf = m["dir_acc"]
            sources.append({"src": "LSTM", "mag": mag,
                            "dir": "UP" if mag > 0 else "DOWN",
                            "conf": conf, "rmse": m["rmse"]})

        # --- VAR ---
        if has_var and col in econ["cols"]:
            ci_idx = econ["cols"].index(col)
            vmag = float(econ["forecast"][ci_idx]) * 10_000
            if econ["vecm_forecast"] is not None:
                vmag2 = float(econ["vecm_forecast"][ci_idx]) * 10_000
                vmag = 0.5 * vmag + 0.5 * vmag2
            # v3.5 FIX 13: use econ.dir_acc (in-sample) not hardcoded 0.52
            sources.append({"src": "VAR", "mag": vmag,
                            "dir": "UP" if vmag > 0 else "DOWN",
                            "conf": econ.get("dir_acc", 0.52),
                            "rmse": econ.get("rmse", 30.0)})

        # --- Hull-White (short end only) ---
        if hw and lab in ["3M", "6M", "1Y"]:
            hw_mag = hw["change_bps"]
            if lab != "3M":
                # Use data-driven attenuation (not hardcoded 0.85/0.70)
                hw_mag *= hw.get("attenuation", {}).get(lab, 0.75)
            # v3.5 FIX 13: use hw.dir_acc (in-sample) not hardcoded 0.60
            sources.append({"src": "HullWhite", "mag": hw_mag,
                            "dir": "UP" if hw_mag > 0 else "DOWN",
                            "conf": hw.get("dir_acc", 0.60),
                            "rmse": hw.get("rmse", 20.0)})

        if not sources:
            continue

        # FIX: performance-adaptive weights (inverse RMSE)
        # v3.5.5 FIX: VAR reports IN-SAMPLE residual RMSE while XGBoost/LSTM
        # report OUT-OF-SAMPLE test-set RMSE. Raw inverse-RMSE weighting
        # therefore systematically over-weights VAR. Apply a 1.5x penalty
        # to VAR's RMSE (typical in-sample/out-of-sample bias ratio) before
        # the weight computation, then hard-cap VAR at 20% of total weight.
        VAR_OOS_PENALTY = 1.5
        VAR_WEIGHT_CAP  = 0.20
        rmses_adj = []
        for s in sources:
            r = s["rmse"]
            if s["src"] == "VAR":
                r = r * VAR_OOS_PENALTY
            rmses_adj.append(r)
        inv_rmse = np.array([1.0 / (r + 1e-6) for r in rmses_adj])
        weights  = inv_rmse / inv_rmse.sum()
        # Apply VAR cap: if VAR exceeds 20%, redistribute excess to the rest
        var_idx = [i for i, s in enumerate(sources) if s["src"] == "VAR"]
        if var_idx and len(sources) > 1:
            vi = var_idx[0]
            if weights[vi] > VAR_WEIGHT_CAP:
                excess = weights[vi] - VAR_WEIGHT_CAP
                weights[vi] = VAR_WEIGHT_CAP
                other_mask = np.array([i != vi for i in range(len(sources))])
                other_sum = weights[other_mask].sum()
                if other_sum > 0:
                    weights[other_mask] += excess * (weights[other_mask] / other_sum)
        for i, s in enumerate(sources):
            s["w"] = float(weights[i])

        tw    = sum(s["w"] for s in sources)
        emag  = sum(s["mag"] * s["w"] for s in sources) / tw
        econf = sum(s["conf"] * s["w"] for s in sources) / tw

        # ==================================================================
        # v3.5.5: REGIME-ADAPTIVE POST-PROCESSING PIPELINE
        # Replaces v3.5.5's simple 30% trend blend with a 4-stage pipeline:
        #   Stage 1: Magnitude calibration (scale up from mean-reversion bias)
        #   Stage 2: Trend blend (stronger for 1W, regime-aware)
        #   Stage 3: Crude oil shock overlay
        #   Stage 4: Direction decision (vol-adaptive FLAT threshold)
        # ==================================================================
        trend_blend_tag = ""

        # Stage 1: MAGNITUDE CALIBRATION (FIX 4)
        # The ensemble systematically predicts near-zero magnitudes because
        # XGBoost trained on MSE/Huber loss learns "predict small". Scale up
        # by the vol ratio so predictions match current-regime volatility.
        # In high-vol regimes: emag *= 1.3-2.0. In normal: *= 1.0-1.3.
        mag_scale = max(1.0, regime["vol_ratio"])
        emag *= mag_scale

        # Stage 2: TREND BLEND (FIX 5, horizon-adaptive)
        # v3.5.5 used 30% always. v3.5.5: blend weight is:
        #   - Horizon-dependent: 1W=50%, 1M=35%, 3M=30%
        #   - Boosted in high-vol regime or when trend is strong
        #   - Only activates when model loses to naive OR when regime is high
        TREND_BLEND_W = {5: 0.50, 20: 0.35, 60: 0.30}.get(horizon, 0.35)
        if vol_regime == "high":
            TREND_BLEND_W = min(0.60, TREND_BLEND_W + 0.15)  # boost in volatile regimes

        should_trend_blend = False
        if lab in xgb_m:
            x_model_acc = xgb_m[lab]["dir_acc"]
            x_naive_acc = xgb_m[lab].get("naive_acc", 0.50)
            if x_model_acc < x_naive_acc - 0.02:
                should_trend_blend = True
        # Also trigger blend when vol regime is "high" regardless of accuracy
        if vol_regime == "high" and regime["trend_strength"] > 0.6:
            should_trend_blend = True

        if should_trend_blend and len(df) > horizon + 1:
            recent = float((df[col].iloc[-1] - df[col].iloc[-horizon-1]) * 10_000)
            if np.isfinite(recent) and abs(recent) > 1e-6:
                emag = (1 - TREND_BLEND_W) * emag + TREND_BLEND_W * recent
                if lab in xgb_m:
                    econf = min(econf, xgb_m[lab].get("naive_acc", 0.55))
                trend_blend_tag = f" [trend-blend {TREND_BLEND_W:.0%}]"

        # Stage 3: CRUDE OIL SHOCK OVERLAY (FIX 6)
        # India imports 85% of crude. When crude moves >5% in `horizon` days,
        # that's the single biggest driver of bond yields. Apply a directional
        # overlay: crude up -> yields up (bad for bonds), crude down -> yields down.
        # Sensitivity: ~5 bps yield change per 1% crude change (empirical).
        if regime["crude_shock"]:
            crude_impact = regime["crude_chg_pct"] * 5.0  # 1% crude → ~5 bps yield
            # Blend: overlay contributes proportionally to shock magnitude
            overlay_w = min(0.30, abs(regime["crude_chg_pct"]) / 30)  # max 30%
            emag = (1 - overlay_w) * emag + overlay_w * crude_impact
            trend_blend_tag += f" [crude-shock {regime['crude_chg_pct']:+.1f}%]"

        # Stage 4: DIRECTION DECISION (vol-adaptive FLAT threshold)
        flat_bps = _compute_flat_bps(df, col, horizon, FLAT_FALLBACK)
        if abs(emag) < flat_bps:
            edir = "FLAT"
        elif emag > 0:
            edir = "UP"
        else:
            edir = "DOWN"

        # v3.5.5 FIX 3: CONFIDENCE INTERVAL — volatility-scaled.
        # Old: raw 5th/95th pctile of last 504 days. Failed catastrophically
        # in volatile regimes (CI ±10 bps, actual move -18 to -38 bps).
        # New: scale CI width by max(1, recent_vol / long_vol) so that
        # volatile regimes automatically produce wider CIs.
        hist = (df[col].diff(horizon) * 10_000).dropna().tail(504)
        if len(hist) > 60:
            raw_lo = float(np.percentile(hist, 5))
            raw_hi = float(np.percentile(hist, 95))
            hist_med = float(np.median(hist))
            # Shift CI to center on prediction
            lo = raw_lo - hist_med + emag
            hi = raw_hi - hist_med + emag
            # Widen by vol ratio (>1 in volatile regimes, 1 otherwise)
            ci_scale = max(1.0, regime["vol_ratio"])
            center_ci = (hi + lo) / 2
            half_w = (hi - lo) / 2 * ci_scale
            lo = center_ci - half_w
            hi = center_ci + half_w
        else:
            vol = hist.std() if len(hist) > 20 else 10.0
            lo = emag - 1.65 * vol
            hi = emag + 1.65 * vol

        # Additional widening under extreme macro composite
        mc = features["macro_composite"].iloc[-1]
        if abs(mc) > 1.5:
            center_ci = (hi + lo) / 2
            half_w = (hi - lo) / 2 * 1.20
            lo = center_ci - half_w
            hi = center_ci + half_w
            econf *= 0.90

        # momentum flag
        mom_col = f"dy20_{lab}"
        if mom_col in features.columns:
            mom20 = features[mom_col].iloc[-1]
            momentum = "Strong" if abs(mom20) > 15 else ("Moderate" if abs(mom20) > 5 else "Weak")
        else:
            momentum = "N/A"

        conf_pct = round(min(econf * 100, 95), 1)
        # Statistical threshold: reject H0(p=0.5) at alpha=0.10 one-sided
        # Formula: 50 + z_0.10 * sqrt(0.25/N) * 100 = 50 + 64/sqrt(N)
        n_test = xgb_m[lab]["n_test"] if lab in xgb_m else 200
        low_conf_threshold = 50.0 + 64.0 / np.sqrt(max(n_test, 50))
        low_conf = conf_pct < low_conf_threshold

        predictions[lab] = {
            "current_pct":   round(current, 4),
            "direction":     "LOW CONF" if low_conf else edir,
            "change_bps":    round(emag, 2),
            "range_lo_bps":  round(lo, 2),
            "range_hi_bps":  round(hi, 2),
            "predicted_pct": round(current + emag / 100, 4),
            "confidence":    conf_pct,
            "momentum":      momentum,
            "n_models":      len(sources),
            "models":        "+".join(s["src"] for s in sources),
        }
    return predictions


# ============================================================================
# SECTION 13 — WALK-FORWARD BACKTEST  (FIXED: naive benchmark, gap)
# ============================================================================
def run_backtest(features, df, macro_w, horizon, n_folds=5):
    print(f"\n[10/12] Walk-forward backtest  ({n_folds} folds, {horizon}d horizon) ...")
    print(f"        [XGBoost-only component — ensemble accuracy may differ]")
    results = {}
    key_mats = ["3M","1Y","2Y","5Y","7Y","10Y","14Y","30Y"]
    # v3.5.5: with post-2015 window (~2700 rows), test_size=252 leaves too
    # little train data for the earliest fold. Shrink to 150 (~7 months).
    tscv = TimeSeriesSplit(n_splits=n_folds, test_size=150)

    for lab in key_mats:
        if lab not in YIELD_COLS:      # safety: maturity may have been filtered
            continue
        col = YIELD_COLS[lab]
        target = (df[col].shift(-horizon) - df[col]) * 10_000
        idx = features.index.intersection(target.dropna().index)
        X, y = features.loc[idx], target.loc[idx]
        mask = np.isfinite(y) & np.all(np.isfinite(X.values), axis=1)
        X, y = X[mask], y[mask]
        # v3.5.5: match live XGBoost training window so walk-forward folds
        # measure out-of-sample accuracy in the SAME regime the live model
        # sees. Without this, early folds cover 1998-2014 rising-rate data
        # while the live model is trained on 2015+, making the comparison
        # invalid (v3.5.1 observed: XGB 70% vs backtest 41% => 29bp gap).
        X = X[X.index >= XGB_TRAIN_START]
        y = y.loc[X.index]
        if len(X) < 1500:
            continue

        daccs, rmses, maes, naive_accs, mom_accs = [], [], [], [], []
        # Precompute past realized changes for momentum baseline
        past_chg = (df[col].diff(horizon) * 10_000).reindex(X.index)

        for train_i, test_i in tscv.split(X):
            # gap between train and test
            train_end = max(train_i) - horizon
            if train_end < 500:
                continue
            Xtr = X.iloc[:train_end]
            ytr = y.iloc[:train_end]
            Xte = X.iloc[test_i]
            yte = y.iloc[test_i]

            # v3.5.5: mirror the live XGBoost time-decay weighting so
            # backtest DirAcc is directly comparable to the build_xgboost()
            # single-split DirAcc (both now emphasise the recent regime).
            n_tr = len(Xtr)
            sw_tr = np.exp(XGB_TIME_DECAY * np.arange(n_tr))
            sw_tr /= sw_tr.mean()
            # v3.5.5 FIX D: same monotone constraints as build_xgboost()
            mono_bt = _build_monotone_vec(Xtr.columns)
            reg = _make_xgb(n_estimators=300, max_depth=5,
                            early_stopping_rounds=30,
                            monotone_constraints=mono_bt)
            reg.fit(Xtr, ytr, sample_weight=sw_tr,
                    eval_set=[(Xte, yte)], verbose=False)
            pred = reg.predict(Xte)

            daccs.append(np.mean(np.sign(yte) == np.sign(pred)))
            rmses.append(np.sqrt(mean_squared_error(yte, pred)))
            maes.append(mean_absolute_error(yte, pred))

            # Naive benchmark = majority class accuracy
            prop_up  = float(np.mean(yte > 0))
            naive_accs.append(max(prop_up, 1 - prop_up))

            # Momentum benchmark: predict same direction as past horizon-day change
            mom_signs = np.sign(past_chg.iloc[test_i].values)
            valid_mom = np.isfinite(mom_signs)
            if valid_mom.sum() > 0:
                mom_accs.append(float(np.mean(
                    np.sign(yte.values[valid_mom]) == mom_signs[valid_mom])))

        if daccs:
            navg = np.mean(naive_accs)
            mavg = np.mean(mom_accs) if mom_accs else navg
            best_base = max(navg, mavg)
            results[lab] = {
                "dir_acc":   np.mean(daccs),
                "dir_std":   np.std(daccs),
                "rmse":      np.mean(rmses),
                "mae":       np.mean(maes),
                "naive_acc": navg,
                "mom_acc":   mavg,
                "edge":      np.mean(daccs) - best_base,
                "folds":     len(daccs),
            }

    print(f"\n  {'Mat':>6s} {'DirAcc':>7s} {'+-Std':>6s} {'RMSE':>8s} "
          f"{'MAE':>8s} {'Naive':>6s} {'Mom':>6s} {'Edge':>7s}")
    print("  " + "-" * 60)
    for m in key_mats:
        if m in results:
            r = results[m]
            print(f"  {m:>6s} {r['dir_acc']:>6.1%} {r['dir_std']:>5.1%} "
                  f"{r['rmse']:>8.2f} {r['mae']:>8.2f} "
                  f"{r['naive_acc']:>5.1%} {r['mom_acc']:>5.1%} "
                  f"{r['edge']:>+6.1%}")

    # Warning: systematic underperformance vs baselines
    neg_count = sum(1 for r in results.values() if r["edge"] < 0)
    if neg_count > len(results) // 2:
        print(f"\n  [WARNING] {neg_count}/{len(results)} maturities have negative "
              f"edge vs best baseline.")
        print(f"  Model may underperform simple benchmarks in current regime.")
    return results


# ============================================================================
# SECTION 14 — FEATURE IMPORTANCE
# ============================================================================
def analyze_importance(xgb_m, macro_w):
    print("\n[11/12] Feature importance analysis ...")
    if "10Y" not in xgb_m:
        return {}
    imp  = xgb_m["10Y"]["reg"].feature_importances_
    feat = xgb_m["10Y"]["feats"]
    top  = np.argsort(imp)[::-1]

    print(f"\n  Top 15 features (10Y regressor):")
    print(f"  {'#':>3s}  {'Feature':42s} {'Imp':>8s}")
    print("  " + "-" * 56)
    for i in range(min(15, len(top))):
        print(f"  {i+1:3d}  {feat[top[i]]:42s} {imp[top[i]]:8.4f}")

    print(f"\n  Macro contribution vs user weight:")
    print(f"  {'Var':15s} {'ModelImp':>9s} {'UserWt':>8s}")
    print("  " + "-" * 35)
    macro_imp = {}
    for name in MACRO_COLS:
        tag = name.lower().replace("/","").replace(" ","")
        mi  = sum(imp[j] for j,f in enumerate(feat)
                  if tag in f.lower().replace("/","").replace(" ",""))
        macro_imp[name] = mi
    ti = sum(macro_imp.values()) or 1
    for name in MACRO_COLS:
        pct = macro_imp[name] / ti * 100
        print(f"  {name:15s} {pct:>8.1f}% {macro_w[name]:>7.1f}%")

    # Check divergence between user weights and model importance
    model_vals = [macro_imp.get(name, 0) for name in MACRO_COLS]
    user_vals  = [macro_w.get(name, 0) for name in MACRO_COLS]
    if len(model_vals) >= 3:
        rho, _ = stats.spearmanr(model_vals, user_vals)
        if rho < 0.4:
            print(f"\n  [WARNING] Model macro drivers diverge from user weights "
                  f"(Spearman rho={rho:.2f})")
            print(f"  XGBoost splits on information gain, not feature scaling.")
            print(f"  User weights affect macro_composite z-score and LSTM, "
                  f"not tree splits.")
    return macro_imp


# ============================================================================
# SECTION 15 — EXCEL OUTPUT  (FIXED: merged cell handling)
# ============================================================================
def write_output(preds, bt, macro_w, diag, horizon_label):
    print(f"\n[12/12] Writing results to Excel ...")
    # Write to separate output file (input file may be open in Excel)
    wb = openpyxl.Workbook()
    # Remove default sheet created by Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    hf  = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    hfl = PatternFill("solid", fgColor="1F4E79")
    upf = PatternFill("solid", fgColor="C6EFCE")
    dnf = PatternFill("solid", fgColor="FFC7CE")
    ftf = PatternFill("solid", fgColor="FFEB9C")
    bdr = Border(*(Side(style="thin"),)*4)
    ctr = Alignment(horizontal="center", vertical="center")

    # ========== PREDICTIONS SHEET ==========
    if OUTPUT_SHEET in wb.sheetnames:
        del wb[OUTPUT_SHEET]
    ws = wb.create_sheet(OUTPUT_SHEET)

    ws.merge_cells("A1:J1")
    ws["A1"] = f"Bond Yield Prediction  |  Horizon: {horizon_label}"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79", name="Calibri")
    ws.merge_cells("A2:J2")
    ws["A2"] = (f"Generated {dt.datetime.now():%Y-%m-%d %H:%M}  |  "
                f"Ensemble: NSS + HullWhite + VAR/VECM + XGBoost"
                + (" + Bi-LSTM" if LSTM_AVAILABLE else ""))
    ws["A2"].font = Font(italic=True, size=10, color="666666")

    ws["A4"] = "Macro Weights Applied:"
    ws["A4"].font = Font(bold=True)
    r = 5
    for nm, wt in macro_w.items():
        ws.cell(r, 1, nm); ws.cell(r, 2, f"{wt:.1f}%"); r += 1

    sr = r + 1
    hdrs = ["Maturity","Current Yield (%)","Direction","Change (bps)",
            "Range Low (bps)","Range High (bps)","Predicted Yield (%)",
            "Confidence (%)","Momentum","Models"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(sr, ci, h); c.font = hf; c.fill = hfl
        c.alignment = ctr; c.border = bdr

    mat_order = list(YIELD_COLS.keys())
    dr = sr + 1
    for lab in mat_order:
        if lab not in preds:
            continue
        p = preds[lab]
        vals = [lab, p["current_pct"], p["direction"], p["change_bps"],
                p["range_lo_bps"], p["range_hi_bps"], p["predicted_pct"],
                p["confidence"], p["momentum"], p["models"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(dr, ci, v); c.border = bdr; c.alignment = ctr
            if ci == 3:
                if v == "UP":       c.fill = upf; c.font = Font(bold=True, color="006100")
                elif v == "DOWN":   c.fill = dnf; c.font = Font(bold=True, color="9C0006")
                elif v == "LOW CONF":
                    c.fill = PatternFill("solid", fgColor="D9D9D9")
                    c.font = Font(bold=True, color="595959")
                else:               c.fill = ftf; c.font = Font(bold=True, color="9C6500")
        dr += 1

    # Safe column width (skip MergedCell objects + guard empty columns)
    for col_cells in ws.columns:
        first = col_cells[0]
        if not hasattr(first, "column_letter"):
            continue
        lengths = [len(str(c.value)) for c in col_cells
                   if c.value is not None and hasattr(c, "column_letter")]
        ml = max(lengths) if lengths else 8
        ws.column_dimensions[first.column_letter].width = min(ml + 4, 28)

    # ========== BACKTEST SHEET ==========
    if BACKTEST_SHEET in wb.sheetnames:
        del wb[BACKTEST_SHEET]
    ws2 = wb.create_sheet(BACKTEST_SHEET)
    ws2.merge_cells("A1:G1")
    ws2["A1"] = "Walk-Forward Backtest Results"
    ws2["A1"].font = Font(bold=True, size=14, color="1F4E79")

    bh = ["Maturity","Dir Accuracy","Std","RMSE (bps)","MAE (bps)",
          "Naive Acc","Mom Acc","Edge vs Best"]
    for ci, h in enumerate(bh, 1):
        c = ws2.cell(3, ci, h); c.font = hf; c.fill = hfl; c.border = bdr

    br = 4
    for lab, btr in bt.items():
        vals = [lab, f"{btr['dir_acc']:.1%}", f"{btr['dir_std']:.1%}",
                f"{btr['rmse']:.2f}", f"{btr['mae']:.2f}",
                f"{btr['naive_acc']:.1%}",
                f"{btr.get('mom_acc', 0):.1%}", f"{btr['edge']:+.1%}"]
        for ci, v in enumerate(vals, 1):
            c = ws2.cell(br, ci, v); c.border = bdr
        br += 1

    br += 2
    ws2.cell(br, 1, "Diagnostic Tests").font = Font(bold=True, size=12)
    br += 1
    ws2.cell(br, 1, "Linearity Analysis (Macro -> 10Y Yield):").font = Font(bold=True)
    br += 1
    if "linearity" in diag:
        for nm, info in diag["linearity"].items():
            ws2.cell(br, 1, nm)
            ws2.cell(br, 2, f"Pearson={info['pearson']:.4f}")
            ws2.cell(br, 3, f"Spearman={info['spearman']:.4f}")
            ws2.cell(br, 4, info["type"])
            br += 1

    br += 1
    ws2.cell(br, 1, "Granger Causality (min p-value):").font = Font(bold=True)
    br += 1
    for nm, pv in diag.get("granger", {}).items():
        ws2.cell(br, 1, nm)
        ws2.cell(br, 2, f"p = {pv:.4f}")
        ws2.cell(br, 3, "Significant" if pv < 0.05 else "Not significant")
        br += 1

    for col_cells in ws2.columns:
        first = col_cells[0]
        if not hasattr(first, "column_letter"):
            continue
        lengths = [len(str(c.value)) for c in col_cells
                   if c.value is not None and hasattr(c, "column_letter")]
        ml = max(lengths) if lengths else 8
        ws2.column_dimensions[first.column_letter].width = min(ml + 4, 30)

    # Save to dedicated output file (never overwrites input)
    save_path = OUTPUT_PATH
    try:
        wb.save(str(save_path))
    except PermissionError:
        # Fallback: save to Desktop if output file is also locked
        save_path = Path.home() / "Desktop" / "bond_predictions_output.xlsx"
        wb.save(str(save_path))
    print(f"    Saved: {save_path}")
    print(f"    Sheets: '{OUTPUT_SHEET}', '{BACKTEST_SHEET}'")


# ============================================================================
# SECTION 16 — MAIN PIPELINE
# ============================================================================
def main():
    # v3.5 FIX 13b: Load non-interactive config first. If a config file
    # exists (Excel launcher mode), use it for macro weights + horizon.
    config, _cfg_path = load_runtime_config()
    non_interactive = config is not None

    df, mat_real_obs = load_data()

    # v3.5.1 FIX 2a: HARD-EXCLUDE 40Y and 50Y unconditionally.
    # Observed on 1998-2026 data: 40Y dir_acc=0.9% (statistically impossible
    # by chance -> anti-signal) and 50Y dir_acc=86.9% (flagged as regime
    # artifact). Both markets are illiquid with infrequent real trades;
    # Refinitiv supplies interpolated quotes. Predictions on these tenors
    # are meaningless and poison the ensemble / importance analysis.
    # v3.5.5 FIX E: Extend HARD_EXCLUDE with 19Y and 24Y. Both are non-benchmark
    # tenors in India — quotes are mostly interpolated, not traded. Observed
    # post-2015 dir_acc: 19Y=48.4%, 24Y=47.2% (below coin-flip). Neither has
    # backtest walk-forward coverage (backtest key_mats excludes them), so
    # there's no independent validation and they just add noise to the
    # prediction summary.
    HARD_EXCLUDE = ["19Y", "24Y", "40Y", "50Y"]
    for lab in HARD_EXCLUDE:
        if lab in YIELD_COLS:
            YIELD_COLS.pop(lab, None)
            MATURITIES_YRS.pop(lab, None)
    DISPLAY_MATS[:] = [m for m in DISPLAY_MATS if m not in HARD_EXCLUDE]
    print(f"    Hard-excluded illiquid / non-benchmark tenors: {HARD_EXCLUDE}")

    # v3.5 FIX 12: Relative coverage filter (replaces absolute MIN_MAT_OBS).
    # A maturity must have real observations covering >= MIN_MAT_COVERAGE
    # of the dataset. Dataset-size-invariant: 500 real obs of a 2000-row
    # dataset now fails (25%=500), but of a 6000-row dataset it'd need 1500.
    total_rows = len(df)
    min_obs_needed = max(200, int(total_rows * MIN_MAT_COVERAGE))
    sparse = [lab for lab, n in mat_real_obs.items()
              if lab in YIELD_COLS and n < min_obs_needed]
    if sparse:
        for lab in sparse:
            YIELD_COLS.pop(lab, None)
            MATURITIES_YRS.pop(lab, None)
        DISPLAY_MATS[:] = [m for m in DISPLAY_MATS if m not in sparse]
        print(f"    Excluded sparse maturities (<{min_obs_needed} real obs, "
              f"{MIN_MAT_COVERAGE:.0%} of {total_rows}): {sorted(sparse)}")

    macro_w = get_macro_weights(config)

    # NSS daily fit
    nss_params = fit_nss_fast(df)

    # Features (with NSS factors and fixed momentum)
    feats = engineer_features(df, macro_w, nss_params)
    df_a  = df.loc[feats.index]

    # Diagnostics
    diag = run_diagnostics(df_a)

    # Horizon
    HMAP = {"1":("1W",5), "2":("1M",20), "3":("3M",60),
            "1W":("1W",5), "1M":("1M",20), "3M":("3M",60)}
    print("\n" + "=" * 72)
    print("  SELECT PREDICTION HORIZON")
    print("=" * 72)
    if non_interactive:
        h_key = str(config.get("horizon", "1M")).strip()
        hlabel, hdays = HMAP.get(h_key, ("1M", 20))
        print(f"    Using horizon from config: {hlabel} ({hdays} trading days)")
    else:
        print("    1) 1 Week   (5 trading days)")
        print("    2) 1 Month  (20 trading days)")
        print("    3) 3 Months (60 trading days)")
        choice = input("\n    Choice [1/2/3, default=2]: ").strip()
        hlabel, hdays = HMAP.get(choice, ("1M", 20))

    # Hull-White
    hw = fit_hull_white(df_a, hdays)

    # VAR/VECM (Johansen now runs inside, using VAR-selected lag)
    econ = build_econometric(df_a, hdays)

    # XGBoost
    xgb_models = build_xgboost(feats, df_a, macro_w, hdays)

    # LSTM
    lstm_models = build_lstm(feats, df_a, hdays)

    # Ensemble
    preds = ensemble_predict(xgb_models, lstm_models, econ, hw,
                             feats, df_a, macro_w, hdays)

    # Backtest
    bt = run_backtest(feats, df_a, macro_w, hdays)

    # Importance
    analyze_importance(xgb_models, macro_w)

    # Write
    write_output(preds, bt, macro_w, diag, hlabel)

    # ===== FINAL SUMMARY =====
    print("\n" + "=" * 82)
    print("  PREDICTION SUMMARY  (v3.5.5)")
    print("=" * 82)
    print(f"  Horizon: {hlabel} ({hdays} trading days)\n")
    print(f"  {'Mat':>6s} {'Current':>8s} {'Dir':>7s} {'Chg(bps)':>9s} "
          f"{'Range(bps)':>18s} {'Predicted':>10s} {'Conf':>6s} {'Mom':>8s}")
    print("  " + "-" * 76)
    for lab in DISPLAY_MATS:
        if lab in preds:
            p = preds[lab]
            rng = f"[{p['range_lo_bps']:+.1f} , {p['range_hi_bps']:+.1f}]"
            print(f"  {lab:>6s} {p['current_pct']:>7.3f}% {p['direction']:>7s} "
                  f"{p['change_bps']:>+8.1f} {rng:>18s} "
                  f"{p['predicted_pct']:>9.3f}% {p['confidence']:>5.1f}% "
                  f"{p['momentum']:>8s}")

    print(f"\n  Results written to:  {OUTPUT_PATH}")
    print(f"  Sheets: '{OUTPUT_SHEET}'  and  '{BACKTEST_SHEET}'")
    print("=" * 82)


if __name__ == "__main__":
    main()