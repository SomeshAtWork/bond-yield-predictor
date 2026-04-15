"""
================================================================================
  BOND YIELD PREDICTOR  -  LAUNCHER BRIDGE
================================================================================
  Called by run_predictor.bat when the user clicks the RUN PREDICTION cell
  in bond_predictor_launcher.xlsx.

  Pipeline:
      1. Read horizon + macro weights from the launcher workbook
      2. Write predictor_config.json (consumed by bond_yield_predictor.py)
      3. Invoke the main predictor in non-interactive mode
      4. Write run status back into the launcher workbook
      5. Open the output workbook for the user

  Why this exists:
      bond_yield_predictor.py runs stand-alone and reads from config JSON
      when present. This bridge is the thin layer that pulls the user's
      most recent inputs out of the launcher .xlsx (which has no macros).
================================================================================
"""

import sys
import json
import time
import subprocess
import datetime as dt
from pathlib import Path

import openpyxl

SCRIPT_DIR    = Path(__file__).resolve().parent
PREDICTOR_PY  = SCRIPT_DIR / "bond_yield_predictor.py"
LAUNCHER_XLSX = SCRIPT_DIR / "bond_predictor_launcher.xlsx"
CONFIG_JSON   = SCRIPT_DIR / "predictor_config.json"
OUTPUT_XLSX   = Path.home() / "Downloads" / "bond_predictions_output.xlsx"


# ---------------------------------------------------------------------------
# Cell references (kept in sync with create_launcher.py)
# ---------------------------------------------------------------------------
WEIGHT_CELLS = {
    "CPI":       "C6",
    "Repo Rate": "C7",
    "IIP":       "C8",
    "USD/INR":   "C9",
    "Crude":     "C10",
    "NSE":       "C11",
    "FII":       "C12",
}
HORIZON_CELL   = "C15"
STATUS_TIME    = "G6"
STATUS_RESULT  = "G7"
STATUS_OUTPUT  = "G8"


def _read_launcher_inputs():
    """Open the launcher .xlsx (read-only), pull user inputs, close.

    We open twice — once read-only to grab inputs (safe even if Excel has
    it open), and later write-mode to update status (requires the workbook
    to be closed in Excel; handled with try/except).
    """
    if not LAUNCHER_XLSX.exists():
        raise FileNotFoundError(
            f"Launcher workbook not found at {LAUNCHER_XLSX}\n"
            f"Run create_launcher.py once to generate it.")

    wb = openpyxl.load_workbook(LAUNCHER_XLSX, read_only=True, data_only=True)
    try:
        ws = wb["Launcher"]
        weights = {}
        for name, addr in WEIGHT_CELLS.items():
            val = ws[addr].value
            try:
                weights[name] = float(val) if val is not None else 10.0
            except (TypeError, ValueError):
                weights[name] = 10.0

        horizon_raw = ws[HORIZON_CELL].value
        horizon = str(horizon_raw).strip().upper() if horizon_raw else "1M"
        if horizon not in ("1W", "1M", "3M"):
            horizon = "1M"

    finally:
        wb.close()
    return weights, horizon


def _write_config(weights, horizon):
    cfg = {
        "horizon": horizon,
        "macro_weights": weights,
        "source": "bond_predictor_launcher.xlsx",
        "written_at": dt.datetime.now().isoformat(timespec="seconds"),
    }
    with open(CONFIG_JSON, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2)
    print(f"  [bridge] Wrote config -> {CONFIG_JSON.name}")


def _run_predictor():
    """Run bond_yield_predictor.py with the current sys.executable.
    Inherits stdin/stdout so the user sees the model progress live.
    """
    if not PREDICTOR_PY.exists():
        raise FileNotFoundError(f"Predictor not found at {PREDICTOR_PY}")

    print(f"  [bridge] Running predictor: {PREDICTOR_PY.name}")
    print("  " + "-" * 68)
    rc = subprocess.call(
        [sys.executable, str(PREDICTOR_PY)],
        cwd=str(SCRIPT_DIR),
    )
    print("  " + "-" * 68)
    print(f"  [bridge] Predictor exited with code {rc}")
    return rc


def _update_status(result_text, output_text):
    """Write status back into the launcher workbook.
    If the workbook is open in Excel, the write will fail — not fatal.
    """
    try:
        wb = openpyxl.load_workbook(LAUNCHER_XLSX)
        ws = wb["Launcher"]
        ws[STATUS_TIME].value   = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws[STATUS_RESULT].value = result_text
        ws[STATUS_OUTPUT].value = output_text
        wb.save(LAUNCHER_XLSX)
        wb.close()
        print(f"  [bridge] Status updated in launcher")
    except PermissionError:
        print(f"  [bridge] Launcher is open in Excel — status not written back")
    except Exception as e:
        print(f"  [bridge] Status update failed: {type(e).__name__}: {e}")


def _open_output():
    """Open the output workbook in the default Excel/WPS handler."""
    if not OUTPUT_XLSX.exists():
        print(f"  [bridge] Output file not found at {OUTPUT_XLSX}")
        return False
    try:
        import os
        os.startfile(str(OUTPUT_XLSX))
        print(f"  [bridge] Opened {OUTPUT_XLSX.name}")
        return True
    except Exception as e:
        print(f"  [bridge] Could not open output: {e}")
        return False


def main():
    print("=" * 72)
    print("  BOND YIELD PREDICTOR  -  LAUNCHER BRIDGE")
    print("=" * 72)
    try:
        weights, horizon = _read_launcher_inputs()
    except Exception as e:
        print(f"ERROR reading launcher: {e}")
        return 3

    total = sum(weights.values()) or 1.0
    print(f"  Horizon: {horizon}")
    print(f"  Weights (raw):")
    for k, v in weights.items():
        pct = v / total * 100
        print(f"    {k:12s}  {v:6.1f}  ({pct:5.1f}% normalised)")

    _write_config(weights, horizon)

    # Delete stale output so we can detect a fresh one
    if OUTPUT_XLSX.exists():
        try:
            OUTPUT_XLSX.unlink()
            print(f"  [bridge] Cleared stale output {OUTPUT_XLSX.name}")
        except PermissionError:
            print(f"  [bridge] WARNING: {OUTPUT_XLSX.name} is open in Excel — "
                  f"close it before running again if you want fresh output")

    t0 = time.time()
    rc = _run_predictor()
    elapsed = time.time() - t0

    if rc == 0 and OUTPUT_XLSX.exists():
        result = f"SUCCESS ({elapsed:.0f}s)"
        print(f"\n  {result}")
        _update_status(result, str(OUTPUT_XLSX))
        _open_output()
        return 0
    else:
        result = f"FAILED (rc={rc}, {elapsed:.0f}s)"
        print(f"\n  {result}")
        _update_status(result, "")
        return rc or 1


if __name__ == "__main__":
    sys.exit(main())
