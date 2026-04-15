"""
================================================================================
  BOND YIELD PREDICTOR  -  EXCEL LAUNCHER BUILDER  (one-time setup)
================================================================================
  Run this ONCE to generate bond_predictor_launcher.xlsx.

  Strategy:
      - No VBA, no macros, no Office security prompts needed.
      - The launcher is a plain .xlsx with input cells (horizon + weights)
        and a styled "button" cell whose HYPERLINK fires run_predictor.bat.
      - The .bat invokes launcher_bridge.py which:
            * reads current cell values from the launcher .xlsx
            * writes predictor_config.json (consumed by bond_yield_predictor.py)
            * runs the predictor in non-interactive mode
            * opens the output workbook

  User workflow after setup:
      1. Open Anaconda (once per session) — only needed if Python isn't on PATH
      2. Double-click bond_predictor_launcher.xlsx
      3. Adjust weights + horizon in the yellow input cells
      4. Click the big RUN PREDICTION hyperlink cell
      5. Console opens, model runs, output appears in Excel automatically

  REQUIREMENTS:
      - Anaconda / Python installed at a standard location
      - openpyxl                (comes with Anaconda by default)
================================================================================
"""

import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run:  pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SCRIPT_DIR    = Path(__file__).resolve().parent
PREDICTOR_PY  = SCRIPT_DIR / "bond_yield_predictor.py"
LAUNCHER_XLSX = SCRIPT_DIR / "bond_predictor_launcher.xlsx"
RUN_BAT       = SCRIPT_DIR / "run_predictor.bat"


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
TITLE_FONT      = Font(bold=True, size=16, color="1F4E79", name="Calibri")
SUBTITLE_FONT   = Font(italic=True, size=10, color="666666", name="Calibri")
SECTION_FONT    = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
SECTION_FILL    = PatternFill("solid", fgColor="1F4E79")
INPUT_LABEL     = Font(bold=True, size=11, name="Calibri")
INPUT_FILL      = PatternFill("solid", fgColor="FFF2CC")     # pale yellow (editable)
LOCKED_FILL     = PatternFill("solid", fgColor="F2F2F2")     # light grey (read-only)
BUTTON_FONT     = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
BUTTON_FILL     = PatternFill("solid", fgColor="2E7D32")     # green
INSTRUCTION     = Font(italic=True, size=10, color="666666", name="Calibri")
NOTE_FONT       = Font(size=9, color="9C6500", name="Calibri")

THIN     = Side(style="thin",  color="808080")
THICK    = Side(style="medium", color="1F4E79")
BORDER_T = Border(left=THIN,  right=THIN,  top=THIN,  bottom=THIN)
BORDER_B = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)
CENTER   = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT     = Alignment(horizontal="left",   vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Build the workbook
# ---------------------------------------------------------------------------
def build_launcher():
    if not PREDICTOR_PY.exists():
        print(f"\nERROR: predictor script not found at {PREDICTOR_PY}")
        sys.exit(1)
    if not RUN_BAT.exists():
        print(f"\nERROR: run_predictor.bat not found at {RUN_BAT}")
        print("       (should have been generated alongside this file)")
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Launcher"

    # ------ Title banner ---------------------------------------------------
    ws.merge_cells("B2:G2")
    ws["B2"] = "Indian Government Bond Yield Predictor"
    ws["B2"].font = TITLE_FONT
    ws["B2"].alignment = CENTER

    ws.merge_cells("B3:G3")
    ws["B3"] = "v3.5  |  VAR/VECM + NSS + Hull-White + XGBoost + Bi-LSTM"
    ws["B3"].font = SUBTITLE_FONT
    ws["B3"].alignment = CENTER

    # ------ Macro weights section -----------------------------------------
    ws.merge_cells("B5:D5")
    ws["B5"] = "  MACRO VARIABLE WEIGHTS"
    ws["B5"].font = SECTION_FONT
    ws["B5"].fill = SECTION_FILL
    ws["B5"].alignment = LEFT

    macro_defaults = [
        ("CPI",       25),
        ("Repo Rate", 25),
        ("IIP",       10),
        ("USD/INR",   10),
        ("Crude",     10),
        ("NSE",       10),
        ("FII",       10),
    ]
    for i, (name, default) in enumerate(macro_defaults):
        row = 6 + i
        ws.cell(row, 2, name).font = INPUT_LABEL
        ws.cell(row, 2).border = BORDER_T

        c = ws.cell(row, 3, default)
        c.fill = INPUT_FILL
        c.font = Font(bold=True, size=11)
        c.alignment = CENTER
        c.border = BORDER_T
        c.number_format = "0"

        ws.cell(row, 4, "%").alignment = LEFT
        ws.cell(row, 4).border = BORDER_T

    # Named cells so launcher_bridge.py can read reliably by name
    # (openpyxl Named Range)
    from openpyxl.workbook.defined_name import DefinedName
    cell_map = {
        "w_cpi":   "Launcher!$C$6",
        "w_repo":  "Launcher!$C$7",
        "w_iip":   "Launcher!$C$8",
        "w_usd":   "Launcher!$C$9",
        "w_crude": "Launcher!$C$10",
        "w_nse":   "Launcher!$C$11",
        "w_fii":   "Launcher!$C$12",
    }
    for nm, ref in cell_map.items():
        wb.defined_names[nm] = DefinedName(name=nm, attr_text=ref)

    # ------ Horizon section -----------------------------------------------
    ws.merge_cells("B14:D14")
    ws["B14"] = "  PREDICTION HORIZON"
    ws["B14"].font = SECTION_FONT
    ws["B14"].fill = SECTION_FILL
    ws["B14"].alignment = LEFT

    ws.cell(15, 2, "Horizon").font = INPUT_LABEL
    ws.cell(15, 2).border = BORDER_T

    c = ws.cell(15, 3, "1M")
    c.fill = INPUT_FILL
    c.font = Font(bold=True, size=11)
    c.alignment = CENTER
    c.border = BORDER_T

    ws.cell(15, 4, "").border = BORDER_T

    # Add data validation (dropdown) for horizon
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"1W,1M,3M"',
                        allow_blank=False, showDropDown=False)
    dv.add("C15")
    ws.add_data_validation(dv)

    wb.defined_names["horizon"] = DefinedName(
        name="horizon", attr_text="Launcher!$C$15")

    ws.cell(16, 2, "(1W = 5d,  1M = 20d,  3M = 60d)").font = NOTE_FONT
    ws.merge_cells("B16:D16")

    # ------ Status section (read-only, populated after each run) ---------
    ws.merge_cells("F5:G5")
    ws["F5"] = "  RUN STATUS"
    ws["F5"].font = SECTION_FONT
    ws["F5"].fill = SECTION_FILL
    ws["F5"].alignment = LEFT

    status_rows = [
        (6, "Last run",   ""),
        (7, "Result",     "Ready"),
        (8, "Output file", ""),
    ]
    for row, label, val in status_rows:
        ws.cell(row, 6, label).font = INPUT_LABEL
        ws.cell(row, 6).border = BORDER_T
        ws.cell(row, 6).fill = LOCKED_FILL
        c = ws.cell(row, 7, val)
        c.font = Font(size=10)
        c.border = BORDER_T
        c.fill = LOCKED_FILL
        c.alignment = LEFT

    wb.defined_names["status_time"]   = DefinedName(name="status_time",   attr_text="Launcher!$G$6")
    wb.defined_names["status_result"] = DefinedName(name="status_result", attr_text="Launcher!$G$7")
    wb.defined_names["status_output"] = DefinedName(name="status_output", attr_text="Launcher!$G$8")

    # ------ THE BUTTON ----------------------------------------------------
    # Styled hyperlink cell that opens run_predictor.bat
    ws.merge_cells("B18:G19")
    btn = ws["B18"]
    btn.value = "►  RUN  PREDICTION  ◄"
    btn.font = BUTTON_FONT
    btn.fill = BUTTON_FILL
    btn.alignment = CENTER
    btn.border = BORDER_B
    btn.hyperlink = str(RUN_BAT)   # absolute path to the .bat

    # Row heights — make the button visually prominent
    ws.row_dimensions[18].height = 28
    ws.row_dimensions[19].height = 28

    # ------ Instructions footer -------------------------------------------
    ws.merge_cells("B21:G21")
    ws["B21"] = "HOW TO USE"
    ws["B21"].font = Font(bold=True, size=11, color="1F4E79")

    instructions = [
        (22, "1.  Adjust weights in yellow cells (C6:C12)  —  total can be any sum; it is auto-normalised."),
        (23, "2.  Set horizon (C15) to 1W, 1M, or 3M."),
        (24, "3.  Click the green RUN PREDICTION bar above."),
        (25, "4.  A console window opens briefly while the model runs (takes 1-3 minutes)."),
        (26, "5.  When complete, bond_predictions_output.xlsx opens automatically with results."),
    ]
    for r, text in instructions:
        ws.merge_cells(f"B{r}:G{r}")
        c = ws.cell(r, 2, text)
        c.font = INSTRUCTION
        c.alignment = LEFT

    ws.merge_cells("B28:G28")
    ws["B28"] = "TIP: The status cells on the right show when the last run completed."
    ws["B28"].font = NOTE_FONT

    ws.merge_cells("B30:G30")
    ws["B30"] = (f"Launcher folder: {SCRIPT_DIR}")
    ws["B30"].font = Font(size=9, color="808080", italic=True)

    # ------ Column widths --------------------------------------------------
    widths = {"A": 2, "B": 18, "C": 12, "D": 4, "E": 3, "F": 14, "G": 26}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ------ Freeze title row -----------------------------------------------
    ws.freeze_panes = "A5"

    # ------ Save -----------------------------------------------------------
    if LAUNCHER_XLSX.exists():
        LAUNCHER_XLSX.unlink()
    wb.save(str(LAUNCHER_XLSX))
    print(f"  OK  ->  {LAUNCHER_XLSX}")


if __name__ == "__main__":
    print("=" * 72)
    print("  BOND YIELD PREDICTOR  -  LAUNCHER BUILDER")
    print("=" * 72)
    build_launcher()
    print()
    print("=" * 72)
    print("  LAUNCHER READY")
    print("=" * 72)
    print(f"  File   : {LAUNCHER_XLSX.name}")
    print(f"  Folder : {SCRIPT_DIR}")
    print()
    print("  Usage:")
    print("    1. Double-click bond_predictor_launcher.xlsx")
    print("    2. Adjust weights + horizon in the yellow cells")
    print("    3. Click the green RUN PREDICTION bar")
    print("    4. Console runs the model and output opens in Excel")
    print("=" * 72)
